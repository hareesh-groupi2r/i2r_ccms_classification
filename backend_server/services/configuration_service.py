"""
Centralized Configuration Service
Manages configuration for all document processing services
"""

import os
import json
import yaml
from pathlib import Path
from typing import Dict, Any, Optional
from dataclasses import dataclass, field
from openpyxl import load_workbook

from .interfaces import IConfigurationService, ProcessingResult, ProcessingStatus


@dataclass
class ServiceConfig:
    """Configuration container for individual services"""
    service_name: str
    enabled: bool = True
    config: Dict[str, Any] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)


class ConfigurationService(IConfigurationService):
    """Centralized configuration management"""
    
    def __init__(self, config_dir: Optional[str] = None):
        self.config_dir = Path(config_dir) if config_dir else Path(__file__).parent.parent
        self.configs: Dict[str, ServiceConfig] = {}
        self._load_configurations()
    
    def _load_configurations(self):
        """Load all service configurations"""
        try:
            # Load default configurations
            self._load_default_configs()
            
            # Load environment-specific overrides
            self._load_environment_configs()
            
            # Load reference data
            self._load_reference_data()
            
        except Exception as e:
            print(f"Warning: Configuration loading failed: {e}")
    
    def _load_default_configs(self):
        """Load default service configurations"""
        
        # Hybrid RAG Classification Service Configuration
        self.configs["hybrid_rag_classification"] = ServiceConfig(
            service_name="hybrid_rag_classification",
            enabled=True,
            config={
                "approach": os.getenv("CLASSIFICATION_APPROACH", "hybrid_rag"),
                "confidence_threshold": float(os.getenv("CLASSIFICATION_CONFIDENCE_THRESHOLD", "0.5")),
                "max_results": int(os.getenv("CLASSIFICATION_MAX_RESULTS", "5")),
                "openai_api_key": os.getenv("OPENAI_API_KEY"),
                "anthropic_api_key": os.getenv("CLAUDE_API_KEY"),
                "gemini_api_key": os.getenv("GEMINI_API_KEY"),
                "include_justification": True,
                "include_issue_types": True,
                "timeout": int(os.getenv("CLASSIFICATION_TIMEOUT", "60"))
            },
            metadata={
                "description": "Hybrid RAG document classification service",
                "version": "1.0.0",
                "required_env_vars": ["OPENAI_API_KEY", "CLAUDE_API_KEY", "GEMINI_API_KEY"]
            }
        )
        
        # Document AI Configuration
        self.configs["docai"] = ServiceConfig(
            service_name="docai",
            config={
                "project_id": os.environ.get("DOCAI_PROJECT_ID", "dynamic-aurora-467007-q5"),
                "location": os.environ.get("DOCAI_LOCATION", "us"),
                "processor_id": os.environ.get("DOCAI_PROCESSOR_ID", "c649821a479ca9b"),
                "timeout": 300,
                "retry_count": 3
            }
        )
        
        # Google Generative AI Configuration
        self.configs["llm"] = ServiceConfig(
            service_name="llm",
            config={
                "api_key": os.environ.get("GOOGLE_API_KEY"),
                "model_name": "gemini-2.0-flash",
                "temperature": 0.1,
                "max_tokens": 4000,
                "timeout": 60,
                "retry_count": 2
            }
        )
        
        # Document Type Classification Configuration
        self.configs["document_type"] = ServiceConfig(
            service_name="document_type",
            config={
                # Correspondence Letters
                "correspondence_keywords": [
                    "letter", "ref", "reference", "subject", "dear sir", "yours faithfully",
                    "yours sincerely", "yours truly", "to whom it may concern", 
                    "attention", "attn", "cc:", "enclosure", "enclosed", "attachment",
                    "memorandum", "memo", "communication", "notice", "circular",
                    "intimation", "reminder", "follow up", "follow-up"
                ],
                "correspondence_patterns": {
                    # Header patterns (high value - strong letter indicators)
                    r"(?i)^(?:to|from|date|subject|ref(?:erence)?)\s*[:.]\s*.+": 0,
                    r"(?i)letter\s*(?:no|ref|id)\s*[:.]\s*\w+": 0,
                    r"(?i)(?:your|our)\s+(?:letter|ref(?:erence)?)\s+(?:no|dated)": 0,
                    
                    # Salutation patterns (medium-high value)
                    r"(?i)dear\s+(?:sir|madam|mr|ms|dr)": 0,
                    r"(?i)to\s+whom\s+it\s+may\s+concern": 0,
                    r"(?i)respected\s+(?:sir|madam)": 0,
                    
                    # Closing patterns (medium value)
                    r"(?i)yours\s+(?:faithfully|sincerely|truly)": 0,
                    r"(?i)(?:kind|warm)\s+regards": 0,
                    r"(?i)thanking\s+you": 0,
                    
                    # Format patterns (medium value)
                    r"(?i)enclos(?:ed?|ure)": 0,
                    r"(?i)(?:cc|copy)\s*[:.]\s*.+": 0,
                    r"(?i)attach(?:ed|ment)": 0,
                    
                    # Standard letter patterns
                    r"(?i)subject\s*[:.]\s*.+": 0,
                    r"(?i)\b(?:ref|reference)\s*[:.]\s*\w+": 0,
                    r"(?i)date\s*[:.]\s*\d{1,2}[\/-]\d{1,2}[\/-]\d{2,4}": 0
                },
                
                # Meeting Minutes (MOMs)
                "meeting_minutes_keywords": [
                    "minutes", "meeting", "mom", "minutes of meeting", "minutes of the meeting",
                    "proceedings", "agenda", "present", "attended by", "attendees",
                    "action items", "action points", "decisions", "resolutions",
                    "next meeting", "follow-up items", "discussions", "deliberations",
                    "chairperson", "secretary", "participants", "venue", "time"
                ],
                "meeting_minutes_patterns": {
                    r"(?i)minutes\s+of\s+(?:the\s+)?meeting": 0,
                    r"(?i)mom\s*[:.#-]": 0,
                    r"(?i)agenda\s+(?:item|point)\s*\d+": 0,
                    r"(?i)present\s*[:.]\s*.+": 0,
                    r"(?i)attended\s+by\s*[:.]\s*.+": 0,
                    r"(?i)action\s+(?:item|point)s?\s*[:.#-]": 0,
                    r"(?i)venue\s*[:.]\s*.+": 0,
                    r"(?i)date\s+&?\s*time\s*[:.]\s*.+": 0
                },
                
                # Progress Reports
                "progress_reports_keywords": [
                    "progress report", "monthly report", "status report", "project report",
                    "completion", "percentage", "milestone", "activities completed",
                    "work done", "achievements", "progress summary", "physical progress",
                    "financial progress", "schedule", "timeline", "delays", "issues faced",
                    "next month", "upcoming activities", "resource utilization"
                ],
                "progress_reports_patterns": {
                    r"(?i)progress\s+report": 0,
                    r"(?i)monthly\s+report": 0,
                    r"(?i)completion\s*[:.]\s*\d+%": 0,
                    r"(?i)physical\s+progress\s*[:.]\s*\d+": 0,
                    r"(?i)financial\s+progress": 0,
                    r"(?i)milestone\s+\d+": 0,
                    r"(?i)activities?\s+completed": 0,
                    r"(?i)work\s+done": 0
                },
                
                # Change Orders/Scope Changes
                "change_orders_keywords": [
                    "change order", "variation", "scope change", "modification",
                    "amendment", "addendum", "revision", "alteration", "extra work",
                    "additional work", "omission", "substitution", "change in scope",
                    "work order", "supplementary agreement", "change directive",
                    "cost impact", "time extension", "schedule impact"
                ],
                "change_orders_patterns": {
                    r"(?i)change\s+order\s*#?\s*\d+": 0,
                    r"(?i)variation\s+order": 0,
                    r"(?i)scope\s+change": 0,
                    r"(?i)modification\s+(?:to|of)\s+contract": 0,
                    r"(?i)amendment\s+(?:to|of)\s+agreement": 0,
                    r"(?i)extra\s+work": 0,
                    r"(?i)additional\s+work": 0,
                    r"(?i)time\s+extension": 0
                },
                
                # Contract Agreements
                "contract_agreements_keywords": [
                    "agreement", "contract", "schedule j", "article 19", "schedule h",
                    "terms and conditions", "scope of work", "specifications",
                    "contractor", "contractee", "work order", "purchase order",
                    "service agreement", "construction contract", "supply contract",
                    "maintenance contract", "tender", "bid", "proposal"
                ],
                "contract_agreements_patterns": {
                    r"(?:This\\s+)?Agreement\\s+is\\s+entered\\s+into": 0,
                    r"(?:\\n|^)\\s*SCHEDULE\\s*[- ]*\\s*J\\s*(?:\\n|$)": "MULTILINE",
                    r"(?:\\n|^)\\s*ARTICLE\\s+19\\s*(?:\\n|$)": "MULTILINE",
                    r"(?:\\n|^)\\s*SCHEDULE\\s*[- ]*\\s*H\\s*(?:\\n|$)": "MULTILINE",
                    r"(?i)terms\s+and\s+conditions": 0,
                    r"(?i)scope\s+of\s+work": 0,
                    r"(?i)contract\s+value\s*[:.]\s*rs\.?\s*\d+": 0,
                    r"(?i)work\s+order\s*#?\s*\w+": 0
                },
                
                # Payment Statements
                "payment_statements_keywords": [
                    "bill", "invoice", "payment", "stage payment statement", "sps",
                    "interim payment certificate", "ipc", "running bill", "final bill",
                    "payment certificate", "work done", "measurement", "quantity",
                    "rate", "amount", "total", "gst", "tax", "deduction", "advance",
                    "retention", "security deposit", "performance guarantee"
                ],
                "payment_statements_patterns": {
                    r"(?i)stage\s+payment\s+statement": 0,
                    r"(?i)interim\s+payment\s+certificate": 0,
                    r"(?i)running\s+bill\s*#?\s*\d+": 0,
                    r"(?i)invoice\s*#?\s*\w+": 0,
                    r"(?i)bill\s*#?\s*\w+": 0,
                    r"(?i)payment\s+certificate": 0,
                    r"(?i)work\s+done\s+up\s+to": 0,
                    r"(?i)total\s+amount\s*[:.]\s*rs\.?\s*[\d,]+": 0
                },
                
                # Court Orders
                "court_orders_keywords": [
                    "court", "order", "judgment", "decree", "writ", "petition",
                    "case", "suit", "hearing", "proceedings", "interim order",
                    "stay order", "injunction", "mandamus", "certiorari",
                    "honorable", "hon'ble", "justice", "judge", "magistrate",
                    "plaintiff", "defendant", "appellant", "respondent"
                ],
                "court_orders_patterns": {
                    r"(?i)in\s+the\s+(?:high\s+)?court\s+of": 0,
                    r"(?i)before\s+the\s+hon'?ble": 0,
                    r"(?i)civil\s+suit\s*#?\s*\d+": 0,
                    r"(?i)w\.?p\.?\s*#?\s*\d+": 0,
                    r"(?i)case\s*#?\s*\w+": 0,
                    r"(?i)order\s+dated\s+\d{1,2}[\/-]\d{1,2}[\/-]\d{2,4}": 0,
                    r"(?i)judgment\s+dated": 0,
                    r"(?i)interim\s+order": 0
                },
                
                # Policy Circulars
                "policy_circulars_keywords": [
                    "circular", "policy", "directive", "guideline", "instruction",
                    "notification", "order", "government order", "go", "memorandum",
                    "advisory", "bulletin", "communique", "announcement",
                    "regulation", "rule", "procedure", "protocol", "standard"
                ],
                "policy_circulars_patterns": {
                    r"(?i)circular\s*#?\s*\w+": 0,
                    r"(?i)government\s+order": 0,
                    r"(?i)notification\s*#?\s*\w+": 0,
                    r"(?i)policy\s+directive": 0,
                    r"(?i)advisory\s*#?\s*\w+": 0,
                    r"(?i)guideline\s+for": 0,
                    r"(?i)instruction\s+dated": 0,
                    r"(?i)memorandum\s*#?\s*\w+": 0
                },
                
                # Technical Drawings
                "technical_drawings_keywords": [
                    "drawing", "plan", "layout", "design", "blueprint", "schematic",
                    "diagram", "sketch", "architectural", "structural", "electrical",
                    "plumbing", "hvac", "section", "elevation", "detail", "assembly",
                    "fabrication", "installation", "construction drawing", "working drawing"
                ],
                "technical_drawings_patterns": {
                    r"(?i)drawing\s*#?\s*\w+": 0,
                    r"(?i)plan\s*#?\s*\w+": 0,
                    r"(?i)sheet\s*#?\s*\d+": 0,
                    r"(?i)scale\s*[:.]\s*1\s*[:]\s*\d+": 0,
                    r"(?i)architectural\s+drawing": 0,
                    r"(?i)structural\s+drawing": 0,
                    r"(?i)working\s+drawing": 0,
                    r"(?i)construction\s+drawing": 0
                },
                
                # General configuration
                "pages_to_check": 1,  # Always 1 page for fast document type classification
                "confidence_threshold": 0.6,
                "min_keyword_matches": 2,
                "pattern_weight": 3,
                "keyword_weight": 1,
                "filename_weight": 2,  # Filename matches are 2× more valuable than keywords
                "length_based_scoring": True,
                
                # Filename-based detection keywords
                "filename_keywords": {
                    "meeting_minutes": ["minutes", "mom", "meeting", "proceedings", "agenda"],
                    "progress_reports": ["progress", "review", "monthly", "status", "report"],
                    "correspondence": ["letter", "correspondence", "memo", "communication"],
                    "contract_agreements": ["agreement", "contract", "tender", "bid"],
                    "payment_statements": ["bill", "invoice", "payment", "sps", "ipc"],
                    "change_orders": ["change", "variation", "modification", "amendment"],
                    "court_orders": ["order", "judgment", "court", "writ", "petition"],
                    "policy_circulars": ["circular", "policy", "notification", "directive"],
                    "technical_drawings": ["drawing", "plan", "layout", "blueprint", "schematic"]
                },
                
                # Document type priority (higher number = higher priority in case of ties)
                # Correspondence letters should be TRUE LAST RESORT (no processing handler for "others")
                # All specific types should be detected first before falling back to correspondence
                "document_type_priority": {
                    "court_orders": 10,           # Legal documents (highest priority)
                    "meeting_minutes": 9,         # Clear structure/agenda patterns
                    "progress_reports": 8,        # Project-specific reports
                    "payment_statements": 7,      # Financial documents
                    "policy_circulars": 6,        # Official announcements
                    "change_orders": 5,           # Project modifications
                    "technical_drawings": 4,      # Drawings/specifications
                    "contract_agreements": 3,     # Large documents (200+ pages)
                    "others": 2,                  # Unprocessable documents
                    "correspondence": 1           # TRUE LAST RESORT (fallback for everything)
                }
            }
        )
        
        # OCR Configuration
        self.configs["ocr"] = ServiceConfig(
            service_name="ocr",
            config={
                "fallback_to_tesseract": True,
                "tesseract_language": "eng",
                "min_text_length": 50,
                "concurrent_pages": 4,
                "image_dpi": 300
            }
        )
        
        # Category Mapping Configuration - XLSX only
        self.configs["category_mapping"] = ServiceConfig(
            service_name="category_mapping",
            config={
                "mapping_file": "issues_to_category_mapping_normalized.xlsx",
                "default_category": "Others",
                "confidence_threshold": 0.7,
                "fuzzy_matching": True
            }
        )
        
        # Normal Document Processing Configuration (Dynamic page limits per document type)
        self.configs["normal_processing"] = ServiceConfig(
            service_name="normal_processing",
            config={
                "processing_method": "standard",
                "text_extraction_method": "comprehensive",
                
                # Document-specific page limits for processing (not classification)
                "pages_per_document_type": {
                    "correspondence": 3,        # 3 pages for correspondence letters
                    "meeting_minutes": 5,       # 5 pages for meeting minutes  
                    "progress_reports": 5,      # 5 pages for progress reports
                    "contract_agreements": -1,  # Full document with smart selection
                    "default": 2                # 2 pages for all other types
                }
            }
        )
        
        # Processing Pipeline Configuration (Full document processing for contracts)
        self.configs["pipeline"] = ServiceConfig(
            service_name="pipeline",
            config={
                "temp_dir_prefix": "ccms_processing_",
                "cleanup_temp_files": True,
                "max_file_size_mb": 200,
                "supported_formats": ["pdf", "png", "jpg", "jpeg", "tiff", "docx"],
                "processing_timeout": 600,
                # PDF Optimization settings
                "optimize_pdfs": True,
                "short_pdf_cache_dir": "/tmp/ccms_short_pdfs",
                "pdf_optimization_keywords": [
                    "agreement", "contract", "scope", "work", "payment", "schedule", 
                    "milestone", "completion", "terms", "conditions", "parties",
                    "contractor", "authority", "value", "amount", "duration"
                ],
                "max_pages_to_analyze": 20,  # For contract agreements - smart page selection
                "always_include_first_pages": 3,
                "always_include_last_pages": 2
            }
        )
    
    def _load_environment_configs(self):
        """Load environment-specific configuration overrides"""
        
        # Check for environment-specific config files
        env_config_file = self.config_dir / "config" / "environment.yaml"
        if env_config_file.exists():
            try:
                with open(env_config_file, 'r') as f:
                    env_configs = yaml.safe_load(f)
                
                for service_name, config_override in env_configs.items():
                    if service_name in self.configs:
                        self.configs[service_name].config.update(config_override)
                        
            except Exception as e:
                print(f"Warning: Failed to load environment config: {e}")
    
    def _load_reference_data(self):
        """Load reference data for services"""
        reference_data = {}
        
        try:
            # Load issue mapping data from XLSX file only
            xlsx_mapping_file = self.config_dir / "issues_to_category_mapping_normalized.xlsx"
            
            if xlsx_mapping_file.exists():
                # Load XLSX using openpyxl
                workbook = load_workbook(xlsx_mapping_file, read_only=True)
                sheet = workbook.active
                
                # Get headers from first row
                headers = [cell.value for cell in sheet[1]]
                
                # Handle different column naming conventions
                issue_type_col = None
                category_col = None
                
                for i, header in enumerate(headers):
                    if header and header.lower().replace('_', '').replace(' ', '') in ['issuetype', 'issuetypes']:
                        issue_type_col = i
                    elif header and header.lower().replace('_', '').replace(' ', '') in ['mappedcategory', 'category']:
                        category_col = i
                
                if issue_type_col is not None and category_col is not None:
                    # Load mappings (skip header row)
                    issue_mappings = {}
                    categories = set()
                    issue_types = set()
                    
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if (len(row) > max(issue_type_col, category_col) and 
                            row[issue_type_col] and row[category_col]):
                            issue_type = str(row[issue_type_col]).strip()
                            category_raw = str(row[category_col]).strip()
                            
                            # Keep original mapping for exact matching
                            issue_mappings[issue_type] = category_raw
                            issue_types.add(issue_type)
                            
                            # Split comma-separated categories and add to unique set
                            if ',' in category_raw:
                                split_categories = [cat.strip() for cat in category_raw.split(',')]
                                categories.update(split_categories)
                            else:
                                categories.add(category_raw)
                    
                    reference_data["issue_mappings"] = issue_mappings
                    reference_data["available_categories"] = sorted(list(categories))
                    reference_data["available_issue_types"] = sorted(list(issue_types))
                    
                workbook.close()
            
            # Load contract field definitions
            contract_fields_file = self.config_dir / "reference_data" / "contract_fields.json"
            if contract_fields_file.exists():
                with open(contract_fields_file, 'r') as f:
                    reference_data["contract_fields"] = json.load(f)
            else:
                # Default contract fields
                reference_data["contract_fields"] = {
                    "Project Name": "Official name of the project",
                    "Contract Value": "Total contract amount with currency",
                    "Contractor Name": "Name of the contracting organization",
                    "Start Date": "Project start date",
                    "End Date": "Project completion date",
                    "Location": "Project location"
                }
            
            # Store reference data in configuration
            self.configs["reference_data"] = ServiceConfig(
                service_name="reference_data",
                config=reference_data
            )
            
        except Exception as e:
            print(f"Warning: Failed to load reference data: {e}")
    
    def get_service_config(self, service_name: str) -> Dict[str, Any]:
        """Get configuration for specific service"""
        if service_name in self.configs:
            return self.configs[service_name].config
        return {}
    
    def get_pages_to_process(self, document_type: str) -> int:
        """Get the number of pages to process for document processing (not classification)"""
        from .interfaces import DocumentType
        
        # Convert string to DocumentType enum if needed
        if isinstance(document_type, str):
            try:
                doc_type = DocumentType(document_type)
            except ValueError:
                doc_type = DocumentType.CORRESPONDENCE  # Default fallback
        else:
            doc_type = document_type
        
        # Get dynamic page limits from normal processing configuration
        processing_config = self.configs["normal_processing"].config
        pages_per_type = processing_config.get("pages_per_document_type", {})
        
        # Convert enum to string if needed
        if hasattr(doc_type, 'value'):
            document_type_key = doc_type.value
        else:
            document_type_key = str(doc_type)
        
        # Return document-specific page count or default
        return pages_per_type.get(document_type_key, pages_per_type.get("default", 2))
    
    
    def update_service_config(self, service_name: str, config: Dict[str, Any]) -> bool:
        """Update configuration for specific service"""
        try:
            if service_name in self.configs:
                self.configs[service_name].config.update(config)
            else:
                self.configs[service_name] = ServiceConfig(
                    service_name=service_name,
                    config=config
                )
            return True
        except Exception as e:
            print(f"Error updating config for {service_name}: {e}")
            return False
    
    def validate_configuration(self) -> ProcessingResult:
        """Validate all service configurations"""
        errors = []
        warnings = []
        
        # Validate LLM configuration
        llm_config = self.get_service_config("llm")
        if not llm_config.get("api_key"):
            errors.append("GOOGLE_API_KEY not configured for LLM service")
        
        # Validate Document AI configuration
        docai_config = self.get_service_config("docai")
        required_docai_fields = ["project_id", "location", "processor_id"]
        for field in required_docai_fields:
            if not docai_config.get(field):
                warnings.append(f"Document AI {field} not configured")
        
        # Validate reference data
        ref_data = self.get_service_config("reference_data")
        if not ref_data.get("issue_mappings"):
            warnings.append("Issue mapping data not loaded")
        
        # Validate file paths
        if not self.config_dir.exists():
            errors.append(f"Configuration directory not found: {self.config_dir}")
        
        status = ProcessingStatus.ERROR if errors else (
            ProcessingStatus.PARTIAL if warnings else ProcessingStatus.SUCCESS
        )
        
        return ProcessingResult(
            status=status,
            data={
                "errors": errors,
                "warnings": warnings,
                "services_configured": list(self.configs.keys())
            }
        )
    
    def get_all_configs(self) -> Dict[str, ServiceConfig]:
        """Get all service configurations"""
        return self.configs.copy()
    
    def is_service_enabled(self, service_name: str) -> bool:
        """Check if a service is enabled"""
        if service_name in self.configs:
            return self.configs[service_name].enabled
        return False
    
    def enable_service(self, service_name: str, enabled: bool = True):
        """Enable or disable a service"""
        if service_name in self.configs:
            self.configs[service_name].enabled = enabled


# Global configuration instance
_config_service = None

def get_config_service() -> ConfigurationService:
    """Get the global configuration service instance"""
    global _config_service
    if _config_service is None:
        _config_service = ConfigurationService()
    return _config_service