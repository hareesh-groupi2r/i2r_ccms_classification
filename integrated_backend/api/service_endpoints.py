"""
Modular Service API Endpoints
Individual REST endpoints for each document processing service
"""

from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename
import os
import tempfile
from pathlib import Path
from typing import Dict, Any

from services.document_type_service import DocumentTypeService
from services.ocr_service import OCRService
from services.llm_service import LLMService
from services.category_mapping_service import CategoryMappingService
from services.document_processing_orchestrator import DocumentProcessingOrchestrator
from services.hybrid_rag_classification_service import get_classification_service
from services.configuration_service import get_config_service
from services.interfaces import ProcessingStatus, DocumentType
import logging

logger = logging.getLogger(__name__)

# Add parent directory to path for batch processor imports
import sys
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# Import batch processor components for ground truth handling
from classifier.category_normalizer import CategoryNormalizer
from classifier.unified_issue_mapper import UnifiedIssueCategoryMapper
from metrics_calculator import MetricsCalculator


# Create Blueprint
service_api = Blueprint('service_api', __name__, url_prefix='/api/services')

# Initialize services
config_service = get_config_service()
document_type_service = DocumentTypeService(config_service)
ocr_service = OCRService(config_service)
llm_service = LLMService(config_service)
category_mapping_service = CategoryMappingService(config_service)
classification_service = get_classification_service(config_service)
orchestrator = DocumentProcessingOrchestrator(config_service)

# Initialize batch processor components for ground truth handling
category_normalizer = CategoryNormalizer(strict_mode=False)
metrics_calculator = MetricsCalculator()

# Initialize unified issue mapper for proper issue→category mapping
project_root = Path(__file__).parent.parent.parent
unified_issue_mapper = UnifiedIssueCategoryMapper(
    training_data_path=str(project_root / "data/raw/Consolidated_labeled_data.xlsx"),
    mapping_file_path=str(project_root / "unified_issue_category_mapping.xlsx")
)

# Helper functions
def allowed_file(filename: str) -> bool:
    """Check if file type is allowed"""
    ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'tiff'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_uploaded_file(file) -> str:
    """Save uploaded file to temporary location"""
    if not file or not allowed_file(file.filename):
        raise ValueError("Invalid file type")
    
    filename = secure_filename(file.filename)
    temp_dir = tempfile.mkdtemp(prefix="ccms_api_")
    temp_path = os.path.join(temp_dir, filename)
    file.save(temp_path)
    return temp_path

def format_processing_result(result) -> Dict[str, Any]:
    """Format ProcessingResult for JSON response"""
    response = {
        "status": result.status.value,
        "success": result.status == ProcessingStatus.SUCCESS
    }
    
    if result.status == ProcessingStatus.SUCCESS:
        response["data"] = result.data
        if result.confidence is not None:
            response["confidence"] = result.confidence
    else:
        response["error"] = result.error_message
    
    if result.metadata:
        response["metadata"] = result.metadata
    
    return response

def cleanup_temp_file(file_path: str):
    """Clean up temporary file and directory"""
    try:
        temp_file = Path(file_path)
        temp_dir = temp_file.parent
        temp_file.unlink(missing_ok=True)
        temp_dir.rmdir()
    except Exception as e:
        print(f"Warning: Could not clean up temp file {file_path}: {e}")

def setup_per_file_logger(pdf_file_name: str, output_folder: str):
    """Setup individual logger for each PDF file"""
    import logging
    import os
    
    # Create safe filename for log
    safe_filename = "".join(c for c in pdf_file_name if c.isalnum() or c in "._- ").rstrip()
    log_filename = f"{safe_filename}.log"
    log_path = os.path.join(output_folder, log_filename)
    
    # Ensure output folder exists
    os.makedirs(output_folder, exist_ok=True)
    
    # Create unique logger name for this file
    logger_name = f"pdf_file_{safe_filename}"
    file_logger = logging.getLogger(logger_name)
    
    # Remove any existing handlers to avoid duplicates
    for handler in file_logger.handlers[:]:
        file_logger.removeHandler(handler)
    
    # Create file handler
    file_handler = logging.FileHandler(log_path, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    
    # Create formatter
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    
    # Add handler to logger
    file_logger.addHandler(file_handler)
    file_logger.setLevel(logging.INFO)
    file_logger.propagate = True  # Still send to main logger
    
    return file_logger, log_path

def cleanup_per_file_logger(file_logger):
    """Clean up file logger handlers"""
    if file_logger:
        for handler in file_logger.handlers[:]:
            handler.close()
            file_logger.removeHandler(handler)

def dual_log(main_logger, file_logger, level, message):
    """Log to both main logger and file logger if file_logger exists"""
    # Always log to main logger
    if level == 'info':
        main_logger.info(message)
    elif level == 'warning':
        main_logger.warning(message)
    elif level == 'error':
        main_logger.error(message)
    elif level == 'debug':
        main_logger.debug(message)
    
    # Also log to file logger if it exists
    if file_logger:
        if level == 'info':
            file_logger.info(message)
        elif level == 'warning':
            file_logger.warning(message)
        elif level == 'error':
            file_logger.error(message)
        elif level == 'debug':
            file_logger.debug(message)

def _auto_detect_ground_truth(pdf_folder: str) -> str:
    """Auto-detect ground truth file in PDF folder - from batch processor"""
    pdf_folder = Path(pdf_folder)
    
    # Common patterns for ground truth files
    patterns = ["EDMS*.xlsx", "LOT-*.xlsx", "*ground_truth*.xlsx", "*_labels.xlsx", "*.xlsx", "*.xls"]
    
    for pattern in patterns:
        ground_truth_files = list(pdf_folder.glob(pattern))
        if ground_truth_files:
            # Prefer files with recognizable patterns first
            for ground_truth_file in ground_truth_files:
                if any(keyword in ground_truth_file.name.lower() 
                      for keyword in ['edms', 'lot', 'ground', 'truth', 'label']):
                    logger.info(f"📊 Auto-detected ground truth: {ground_truth_file.name}")
                    return str(ground_truth_file)
            
            # If no keyword match, use first file
            logger.info(f"📊 Using Excel file as ground truth: {ground_truth_files[0].name}")
            return str(ground_truth_files[0])
    
    logger.info("📊 No ground truth file auto-detected")
    return None

def _load_ground_truth(ground_truth_file: str) -> Dict:
    """Load ground truth from Excel file - complete batch processor version"""
    import pandas as pd
    try:
        df = pd.read_excel(ground_truth_file)
        
        # Auto-detect the format and find the correct columns
        ground_truth = {}
        
        # Check if this looks like the LOT-21 format (has "Sr. No" in first data row)
        if len(df) > 1 and str(df.iloc[1, 0]).strip() == "Sr. No":
            logger.info("📊 Detected LOT-21 ground truth format")
            # LOT-21 format: skip first 2 rows (headers), file name in column 2, categories in column 5
            
            for idx in range(2, len(df)):
                row = df.iloc[idx]
                
                # File name is in column 2, add .pdf extension if missing
                file_name_raw = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ""
                if not file_name_raw or file_name_raw.strip() in ["", "nan"]:
                    continue
                
                file_name = file_name_raw.strip()
                if not file_name.lower().endswith('.pdf'):
                    file_name = file_name + '.pdf'
                
                # Categories are in column 5, comma-separated
                categories_raw = str(row.iloc[5]) if pd.notna(row.iloc[5]) else ""
                
                if categories_raw and categories_raw.strip() not in ["", "nan"]:
                    # Use the CategoryNormalizer to parse and normalize categories
                    normalized_categories = category_normalizer.parse_and_normalize_categories(categories_raw)
                    
                    # Consolidate categories for the same file
                    if file_name in ground_truth:
                        # Add new categories to existing ones, avoiding duplicates
                        for cat in normalized_categories:
                            if cat not in ground_truth[file_name]:
                                ground_truth[file_name].append(cat)
                    else:
                        # First occurrence of this file
                        ground_truth[file_name] = normalized_categories

        else:
            # Generic format: first column is file name, remaining columns are categories  
            logger.info("📊 Using generic ground truth format")
            for _, row in df.iterrows():
                file_name = str(row.iloc[0])  # First column is file name
                if not file_name or file_name.strip() in ["", "nan"]:
                    continue
                    
                raw_categories = []
                # Collect non-empty categories from remaining columns
                for col_idx in range(1, len(row)):
                    if pd.notna(row.iloc[col_idx]) and str(row.iloc[col_idx]).strip():
                        cat = str(row.iloc[col_idx]).strip()
                        if cat not in ["", "nan"]:
                            raw_categories.append(cat)
                
                # Normalize categories using CategoryNormalizer
                normalized_categories = []
                for cat in raw_categories:
                    norm_cat, status, confidence = category_normalizer.normalize_category(cat)
                    if norm_cat and norm_cat not in normalized_categories:
                        normalized_categories.append(norm_cat)
                
                # Consolidate categories for the same file
                if file_name in ground_truth:
                    # Add new categories to existing ones, avoiding duplicates
                    for cat in normalized_categories:
                        if cat not in ground_truth[file_name]:
                            ground_truth[file_name].append(cat)
                else:
                    # First occurrence of this file
                    ground_truth[file_name] = normalized_categories
        
        # Remove files with no categories
        ground_truth = {k: v for k, v in ground_truth.items() if v}
        
        logger.info(f"📊 Loaded ground truth for {len(ground_truth)} files")
        if logger.isEnabledFor(logging.DEBUG):
            for file_name, categories in list(ground_truth.items())[:5]:  # Show first 5
                logger.debug(f"📊 {file_name}: {categories}")
        
        return ground_truth
        
    except Exception as e:
        logger.error(f"Failed to load ground truth from {ground_truth_file}: {e}")
        return {}


# Document Type Classification Endpoints
@service_api.route('/document-type/classify', methods=['POST'])
def classify_document_type():
    """Classify document type from uploaded file"""
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Save uploaded file
        temp_path = save_uploaded_file(file)
        
        try:
            # Get options from request
            options = {
                "pages_to_check": request.form.get("pages_to_check", type=int),
                "use_advanced_classification": request.form.get("use_advanced_classification", "true").lower() == "true",
                "confidence_threshold": request.form.get("confidence_threshold", type=float)
            }
            # Remove None values
            options = {k: v for k, v in options.items() if v is not None}
            
            # Classify document
            result = document_type_service.classify_document(temp_path, **options)
            
            # Format response
            response = format_processing_result(result)
            if result.status == ProcessingStatus.SUCCESS and isinstance(result.data, DocumentType):
                response["data"] = result.data.value
            
            return jsonify(response)
            
        finally:
            cleanup_temp_file(temp_path)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/document-type/classify-text', methods=['POST'])
def classify_document_type_from_text():
    """Classify document type from text content"""
    try:
        data = request.get_json()
        if not data or 'text' not in data:
            return jsonify({"error": "No text content provided"}), 400
        
        text_content = data['text']
        options = {
            "use_advanced_classification": data.get("use_advanced_classification", True),
            "confidence_threshold": data.get("confidence_threshold")
        }
        # Remove None values
        options = {k: v for k, v in options.items() if v is not None}
        
        result = document_type_service.classify_from_text(text_content, **options)
        
        response = format_processing_result(result)
        if result.status == ProcessingStatus.SUCCESS and isinstance(result.data, DocumentType):
            response["data"] = result.data.value
        
        return jsonify(response)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# OCR Service Endpoints
@service_api.route('/ocr/extract', methods=['POST'])
def extract_text_ocr():
    """Extract text from uploaded document"""
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Save uploaded file
        temp_path = save_uploaded_file(file)
        
        try:
            # Get options from request
            options = {
                "method": request.form.get("method", "auto"),
                "fallback_on_error": request.form.get("fallback_on_error", "true").lower() == "true"
            }
            
            # Handle page range
            page_range = request.form.get("page_range")
            if page_range:
                try:
                    start, end = map(int, page_range.split("-"))
                    options["page_range"] = (start, end)
                except ValueError:
                    return jsonify({"error": "Invalid page_range format. Use 'start-end' (e.g., '1-5')"}), 400
            
            result = ocr_service.extract_text(temp_path, **options)
            return jsonify(format_processing_result(result))
            
        finally:
            cleanup_temp_file(temp_path)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/ocr/extract-pages', methods=['POST'])
def extract_text_from_pages():
    """Extract text from specific pages"""
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Get page numbers
        page_numbers = request.form.get("page_numbers")
        if not page_numbers:
            return jsonify({"error": "No page numbers provided"}), 400
        
        try:
            page_numbers = [int(p.strip()) for p in page_numbers.split(",")]
        except ValueError:
            return jsonify({"error": "Invalid page numbers format. Use comma-separated integers (e.g., '1,3,5')"}), 400
        
        # Save uploaded file
        temp_path = save_uploaded_file(file)
        
        try:
            options = {
                "method": request.form.get("method", "pypdf")
            }
            
            result = ocr_service.extract_text_from_pages(temp_path, page_numbers, **options)
            return jsonify(format_processing_result(result))
            
        finally:
            cleanup_temp_file(temp_path)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/ocr/methods', methods=['GET'])
def get_available_ocr_methods():
    """Get list of available OCR methods"""
    try:
        result = ocr_service.get_available_methods()
        return jsonify(format_processing_result(result))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# LLM Service Endpoints
@service_api.route('/llm/extract-structured', methods=['POST'])
def extract_structured_data():
    """Extract structured data from text using LLM"""
    try:
        data = request.get_json()
        if not data or 'text' not in data or 'schema' not in data:
            return jsonify({"error": "Both 'text' and 'schema' are required"}), 400
        
        text_content = data['text']
        extraction_schema = data['schema']
        
        options = {
            "output_format": data.get("output_format", "json"),
            "additional_instructions": data.get("additional_instructions"),
            "max_retries": data.get("max_retries")
        }
        # Remove None values
        options = {k: v for k, v in options.items() if v is not None}
        
        result = llm_service.extract_structured_data(text_content, extraction_schema, **options)
        return jsonify(format_processing_result(result))
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/llm/classify-content', methods=['POST'])
def classify_content():
    """Classify content using LLM"""
    try:
        data = request.get_json()
        if not data or 'text' not in data or 'options' not in data:
            return jsonify({"error": "Both 'text' and 'options' are required"}), 400
        
        text_content = data['text']
        classification_options = data['options']
        
        kwargs = {
            "confidence_threshold": data.get("confidence_threshold"),
            "additional_context": data.get("additional_context"),
            "return_all_scores": data.get("return_all_scores", False),
            "max_retries": data.get("max_retries")
        }
        # Remove None values
        kwargs = {k: v for k, v in kwargs.items() if v is not None}
        
        result = llm_service.classify_content(text_content, classification_options, **kwargs)
        return jsonify(format_processing_result(result))
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/llm/status', methods=['GET'])
def get_llm_service_status():
    """Get LLM service status"""
    try:
        result = llm_service.get_service_status()
        return jsonify(format_processing_result(result))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Category Mapping Service Endpoints
@service_api.route('/category-mapping/map-issue', methods=['POST'])
def map_issue_to_category():
    """Map issue type to category"""
    try:
        data = request.get_json()
        if not data or 'issue_type' not in data:
            return jsonify({"error": "issue_type is required"}), 400
        
        issue_type = data['issue_type']
        options = {
            "use_fuzzy_matching": data.get("use_fuzzy_matching"),
            "min_confidence": data.get("min_confidence"),
            "case_sensitive": data.get("case_sensitive", False)
        }
        # Remove None values
        options = {k: v for k, v in options.items() if v is not None}
        
        result = category_mapping_service.map_issue_to_category(issue_type, **options)
        return jsonify(format_processing_result(result))
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/category-mapping/bulk-classify', methods=['POST'])
def bulk_classify_issues():
    """Classify multiple issues at once"""
    try:
        data = request.get_json()
        if not data or 'issues' not in data:
            return jsonify({"error": "issues list is required"}), 400
        
        issues = data['issues']
        if not isinstance(issues, list):
            return jsonify({"error": "issues must be a list"}), 400
        
        options = {
            "use_fuzzy_matching": data.get("use_fuzzy_matching"),
            "min_confidence": data.get("min_confidence"),
            "case_sensitive": data.get("case_sensitive", False)
        }
        # Remove None values
        options = {k: v for k, v in options.items() if v is not None}
        
        result = category_mapping_service.bulk_classify_issues(issues, **options)
        return jsonify(format_processing_result(result))
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/category-mapping/categories', methods=['GET'])
def get_available_categories():
    """Get list of available categories"""
    try:
        include_counts = request.args.get('include_counts', 'false').lower() == 'true'
        sort_by = request.args.get('sort_by', 'name')
        
        result = category_mapping_service.get_available_categories(
            include_counts=include_counts,
            sort_by=sort_by
        )
        return jsonify(format_processing_result(result))
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/category-mapping/categories/<category>/issues', methods=['GET'])
def get_issues_for_category(category: str):
    """Get issue types for specific category"""
    try:
        case_sensitive = request.args.get('case_sensitive', 'false').lower() == 'true'
        
        result = category_mapping_service.get_issue_types_for_category(
            category, 
            case_sensitive=case_sensitive
        )
        return jsonify(format_processing_result(result))
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/category-mapping/statistics', methods=['GET'])
def get_mapping_statistics():
    """Get mapping statistics"""
    try:
        result = category_mapping_service.get_mapping_statistics()
        return jsonify(format_processing_result(result))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/category-mapping/add-mapping', methods=['POST'])
def add_issue_mapping():
    """Add new issue to category mapping"""
    try:
        data = request.get_json()
        if not data or 'issue_type' not in data or 'category' not in data:
            return jsonify({"error": "Both issue_type and category are required"}), 400
        
        issue_type = data['issue_type']
        category = data['category']
        save_to_file = data.get('save_to_file', False)
        
        result = category_mapping_service.add_mapping(issue_type, category, save_to_file=save_to_file)
        return jsonify(format_processing_result(result))
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/category-mapping/remove-mapping', methods=['DELETE'])
def remove_issue_mapping():
    """Remove issue to category mapping"""
    try:
        data = request.get_json()
        if not data or 'issue_type' not in data:
            return jsonify({"error": "issue_type is required"}), 400
        
        issue_type = data['issue_type']
        save_to_file = data.get('save_to_file', False)
        
        result = category_mapping_service.remove_mapping(issue_type, save_to_file=save_to_file)
        return jsonify(format_processing_result(result))
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Document Processing Orchestrator Endpoints
@service_api.route('/orchestrator/process-document', methods=['POST'])
def process_document_full():
    """Process document end-to-end with all services"""
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Save uploaded file
        temp_path = save_uploaded_file(file)
        
        try:
            # Parse processing options
            processing_options = {}
            
            # Boolean options
            bool_options = ["extract_structured_data", "classify_issues", "use_advanced_classification", "fallback_on_error"]
            for option in bool_options:
                value = request.form.get(option)
                if value is not None:
                    processing_options[option] = value.lower() == "true"
            
            # String options
            string_options = ["document_type", "extraction_method", "output_format", "llm_instructions"]
            for option in string_options:
                value = request.form.get(option)
                if value:
                    processing_options[option] = value
            
            # Numeric options
            numeric_options = ["pages_to_check", "llm_retries"]
            for option in numeric_options:
                value = request.form.get(option, type=int)
                if value is not None:
                    processing_options[option] = value
            
            # Float options
            float_options = ["classification_confidence", "confidence_threshold"]
            for option in float_options:
                value = request.form.get(option, type=float)
                if value is not None:
                    processing_options[option] = value
            
            # Custom extraction schema (JSON string)
            schema_str = request.form.get("extraction_schema")
            if schema_str:
                try:
                    import json
                    processing_options["extraction_schema"] = json.loads(schema_str)
                except json.JSONDecodeError:
                    return jsonify({"error": "Invalid JSON in extraction_schema"}), 400
            
            result = orchestrator.process_document(temp_path, processing_options)
            
            # Format response - handle complex objects
            response = format_processing_result(result)
            if result.status == ProcessingStatus.SUCCESS and hasattr(result.data, '__dict__'):
                # Convert dataclass to dict
                data_dict = {}
                for key, value in result.data.__dict__.items():
                    if isinstance(value, DocumentType):
                        data_dict[key] = value.value
                    elif hasattr(value, '__dict__'):
                        data_dict[key] = value.__dict__
                    else:
                        data_dict[key] = value
                response["data"] = data_dict
            
            return jsonify(response)
            
        finally:
            cleanup_temp_file(temp_path)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/orchestrator/process-partial', methods=['POST'])
def process_document_partial():
    """Process document with specific services only"""
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Get services to use
        services_str = request.form.get("services")
        if not services_str:
            return jsonify({"error": "services parameter is required"}), 400
        
        services = [s.strip() for s in services_str.split(",")]
        valid_services = ["document_type", "ocr", "llm", "category_mapping"]
        
        for service in services:
            if service not in valid_services:
                return jsonify({"error": f"Invalid service: {service}. Valid options: {valid_services}"}), 400
        
        # Save uploaded file
        temp_path = save_uploaded_file(file)
        
        try:
            # Parse options
            kwargs = {}
            
            # For LLM service
            schema_str = request.form.get("extraction_schema")
            if schema_str and "llm" in services:
                try:
                    import json
                    kwargs["extraction_schema"] = json.loads(schema_str)
                except json.JSONDecodeError:
                    return jsonify({"error": "Invalid JSON in extraction_schema"}), 400
            
            # For category mapping
            issues_str = request.form.get("issues")
            if issues_str and "category_mapping" in services:
                kwargs["issues"] = [i.strip() for i in issues_str.split(",")]
            
            # Other options
            for key in ["method", "document_type", "fuzzy_matching"]:
                value = request.form.get(key)
                if value:
                    kwargs[key] = value
            
            result = orchestrator.process_document_partial(temp_path, services, **kwargs)
            
            # Format response - handle ProcessingResult objects in data
            response = format_processing_result(result)
            if result.status == ProcessingStatus.SUCCESS and isinstance(result.data, dict):
                formatted_data = {}
                for key, value in result.data.items():
                    if hasattr(value, 'status'):  # It's a ProcessingResult
                        formatted_data[key] = format_processing_result(value)
                        # Handle document type enum
                        if key == "document_type" and value.status == ProcessingStatus.SUCCESS:
                            if isinstance(value.data, DocumentType):
                                formatted_data[key]["data"] = value.data.value
                    else:
                        formatted_data[key] = value
                response["data"] = formatted_data
            
            return jsonify(response)
            
        finally:
            cleanup_temp_file(temp_path)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/orchestrator/status', methods=['GET'])
def get_orchestrator_status():
    """Get orchestrator and all services status"""
    try:
        result = orchestrator.get_orchestrator_status()
        return jsonify(format_processing_result(result))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Configuration Service Endpoints
@service_api.route('/config/validate', methods=['GET'])
def validate_configuration():
    """Validate all service configurations"""
    try:
        result = config_service.validate_configuration()
        return jsonify(format_processing_result(result))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/config/services', methods=['GET'])
def get_all_service_configs():
    """Get all service configurations"""
    try:
        configs = config_service.get_all_configs()
        
        # Convert ServiceConfig objects to dicts
        configs_dict = {}
        for name, config in configs.items():
            configs_dict[name] = {
                "service_name": config.service_name,
                "enabled": config.enabled,
                "config": config.config,
                "metadata": config.metadata
            }
        
        return jsonify({
            "status": "success",
            "success": True,
            "data": configs_dict
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/config/services/<service_name>', methods=['GET'])
def get_service_config(service_name: str):
    """Get configuration for specific service"""
    try:
        config = config_service.get_service_config(service_name)
        if not config:
            return jsonify({"error": f"Service '{service_name}' not found"}), 404
        
        return jsonify({
            "status": "success",
            "success": True,
            "data": config
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/config/services/<service_name>', methods=['PUT'])
def update_service_config(service_name: str):
    """Update configuration for specific service"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No configuration data provided"}), 400
        
        success = config_service.update_service_config(service_name, data)
        
        return jsonify({
            "status": "success" if success else "error",
            "success": success,
            "message": f"Configuration updated for {service_name}" if success else f"Failed to update configuration for {service_name}"
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Health Check Endpoint
@service_api.route('/health', methods=['GET'])
def health_check():
    """Overall health check for all services"""
    try:
        status_result = orchestrator.get_orchestrator_status()
        
        if status_result.status == ProcessingStatus.SUCCESS:
            all_healthy = True
            for service_data in status_result.data.get("services", {}).values():
                if not service_data.get("available", True):
                    all_healthy = False
                    break
            
            return jsonify({
                "status": "healthy" if all_healthy else "degraded",
                "timestamp": datetime.utcnow().isoformat(),
                "services": status_result.data["services"]
            }), 200 if all_healthy else 206
        
        else:
            return jsonify({
                "status": "unhealthy",
                "error": status_result.error_message,
                "timestamp": datetime.utcnow().isoformat()
            }), 503
    
    except Exception as e:
        return jsonify({
            "status": "error",
            "error": str(e),
            "timestamp": datetime.utcnow().isoformat()
        }), 500


# Hybrid RAG Classification Service Endpoints
@service_api.route('/hybrid-rag-classification/classify-document', methods=['POST'])
def classify_document_by_id():
    """Classify document by fetching content from database using document ID"""
    try:
        data = request.get_json()
        if not data or 'document_id' not in data:
            return jsonify({"error": "document_id is required"}), 400
        
        document_id = data['document_id']
        options = data.get('options', {})
        
        result = classification_service.classify_document_by_id(document_id, **options)
        return jsonify(format_processing_result(result))
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/hybrid-rag-classification/classify-text', methods=['POST'])
def classify_text_content():
    """Classify text content directly (subject and body)"""
    try:
        data = request.get_json()
        if not data or 'subject' not in data:
            return jsonify({"error": "subject is required"}), 400
        
        subject = data['subject']
        body = data.get('body', '')
        options = data.get('options', {})
        
        result = classification_service.classify_text(subject, body, **options)
        return jsonify(format_processing_result(result))
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/hybrid-rag-classification/classify-batch', methods=['POST'])
def classify_batch_texts():
    """Classify multiple text documents in batch"""
    try:
        data = request.get_json()
        if not data or 'texts' not in data:
            return jsonify({"error": "texts array is required"}), 400
        
        texts = data['texts']
        if not isinstance(texts, list):
            return jsonify({"error": "texts must be an array"}), 400
        
        if len(texts) == 0:
            return jsonify({"error": "texts array cannot be empty"}), 400
        
        if len(texts) > 50:  # Limit batch size
            return jsonify({"error": "Maximum batch size is 50 documents"}), 400
        
        options = data.get('options', {})
        
        result = classification_service.classify_batch(texts, **options)
        return jsonify(format_processing_result(result))
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/hybrid-rag-classification/categories', methods=['GET'])
def get_hybrid_rag_categories():
    """Get list of available categories for classification"""
    try:
        result = classification_service.get_available_categories()
        return jsonify(format_processing_result(result))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/hybrid-rag-classification/issue-types', methods=['GET'])
def get_available_issue_types():
    """Get list of available issue types for classification"""
    try:
        result = classification_service.get_available_issue_types()
        return jsonify(format_processing_result(result))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_api.route('/hybrid-rag-classification/status', methods=['GET'])
def get_classification_service_status():
    """Get classification service status and statistics"""
    try:
        result = classification_service.get_service_status()
        return jsonify(format_processing_result(result))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Import datetime for health check
from datetime import datetime
import pandas as pd


# Batch Processing Endpoints for Folder Processing
@service_api.route('/hybrid-rag-classification/process-folder', methods=['POST'])
def process_pdf_folder():
    """Process a folder of PDF files with batch processing capabilities"""
    try:
        data = request.get_json()
        
        # Support both parameter names for backwards compatibility
        pdf_folder = data.get('pdf_folder') or data.get('folder_path')
        if not pdf_folder:
            return jsonify({"error": "pdf_folder or folder_path is required"}), 400
        
        # Validate folder exists
        if not Path(pdf_folder).exists():
            return jsonify({"error": f"Folder not found: {pdf_folder}"}), 404
        
        # Count PDF files
        pdf_files = list(Path(pdf_folder).glob("*.pdf"))
        if len(pdf_files) == 0:
            return jsonify({"error": "No PDF files found in folder"}), 400
        
        # Get options
        options = data.get('options', {})
        ground_truth_file = data.get('ground_truth_file')
        output_folder = data.get('output_folder', './results')
        enable_metrics = data.get('enable_metrics', True)
        include_patterns = data.get('include_patterns', [])
        max_files = data.get('max_files', None)
        
        # Auto-detect ground truth Excel file using batch processor logic
        if not ground_truth_file and enable_metrics:
            ground_truth_file = _auto_detect_ground_truth(pdf_folder)
            if ground_truth_file:
                logger.info(f"📊 Auto-detected ground truth file: {Path(ground_truth_file).name}")
            else:
                logger.info("📊 No ground truth file auto-detected")
        
        # Load ground truth if available using batch processor logic
        ground_truth_data = {}
        if ground_truth_file and enable_metrics:
            ground_truth_data = _load_ground_truth(ground_truth_file)
            logger.info(f"📊 Loaded ground truth for {len(ground_truth_data)} files")
        
        # Enhanced logging options
        debug_logging = data.get('debug_logging', False)
        per_file_logging = data.get('per_file_logging', False)
        log_text_extraction = data.get('log_text_extraction', False)
        log_classification_details = data.get('log_classification_details', False)
        create_per_file_logs = data.get('create_per_file_logs', False) or per_file_logging
        
        # Process each PDF using the already-initialized classification service
        import time
        from datetime import datetime
        import fnmatch
        
        # Get all PDF files
        all_pdf_files = list(Path(pdf_folder).glob("*.pdf"))
        
        # Filter by include_patterns if specified
        if include_patterns:
            pdf_files = []
            for pdf_file in all_pdf_files:
                for pattern in include_patterns:
                    if fnmatch.fnmatch(pdf_file.name, pattern):
                        pdf_files.append(pdf_file)
                        break  # Only add once even if multiple patterns match
            
            if len(pdf_files) == 0:
                return jsonify({"error": f"No PDF files match patterns: {include_patterns}"}), 400
                
            logger.info(f"Filtered {len(all_pdf_files)} files to {len(pdf_files)} files using patterns: {include_patterns}")
        else:
            pdf_files = all_pdf_files
        
        # Apply max_files limit if specified
        if max_files and max_files > 0 and len(pdf_files) > max_files:
            original_count = len(pdf_files)
            pdf_files = pdf_files[:max_files]
            logger.info(f"Limited to {max_files} files (was {original_count} files)")
        
        # Initialize results structure
        results = {
            'processing_stats': {
                'total_files': len(pdf_files),
                'processed_files': 0,
                'failed_files': 0,
                'start_time': datetime.now(),
                'end_time': None
            },
            'results': [],
            'overall_metrics': {}
        }
        
        # Set up enhanced logging if requested
        if debug_logging:
            logging.getLogger().setLevel(logging.DEBUG)
            logging.getLogger('classifier').setLevel(logging.DEBUG)
            
        # Process each PDF file
        for pdf_file in pdf_files:
            file_logger = None
            try:
                processing_start = time.time()
                
                # Setup per-file logger if requested
                if create_per_file_logs:
                    file_logger, log_path = setup_per_file_logger(pdf_file.name, output_folder)
                    file_logger.info(f"📄 PROCESSING FILE: {pdf_file.name}")
                    file_logger.info(f"📁 File Path: {pdf_file}")
                    file_logger.info(f"📊 File Size: {pdf_file.stat().st_size:,} bytes")
                    file_logger.info(f"{'='*80}")
                    logger.info(f"📄 Processing {pdf_file.name} (log: {log_path})")
                elif per_file_logging:
                    logger.info(f"\n{'='*80}")
                    logger.info(f"📄 PROCESSING FILE: {pdf_file.name}")
                    logger.info(f"📁 File Path: {pdf_file}")
                    logger.info(f"📊 File Size: {pdf_file.stat().st_size:,} bytes")
                    logger.info(f"{'='*80}")
                else:
                    logger.info(f"Processing {pdf_file.name}...")
                
                # Extract text from PDF using same method as batch processor
                if log_text_extraction:
                    dual_log(logger, file_logger, 'info', f"🔍 Starting text extraction...")
                    
                from classifier.pdf_extractor import PDFExtractor
                pdf_extractor = PDFExtractor(max_pages=2)  # Same as batch processor
                
                text_start = time.time()
                extracted_text, extraction_method = pdf_extractor.extract_text(str(pdf_file))
                text_time = time.time() - text_start
                
                if log_text_extraction:
                    dual_log(logger, file_logger, 'info', f"✅ Text extraction completed:")
                    dual_log(logger, file_logger, 'info', f"   📏 Characters extracted: {len(extracted_text):,}")
                    dual_log(logger, file_logger, 'info', f"   ⚡ Extraction method: {extraction_method}")
                    dual_log(logger, file_logger, 'info', f"   ⏱️  Extraction time: {text_time:.2f}s")
                    if len(extracted_text) > 0:
                        preview = extracted_text[:200].replace('\n', ' ')
                        dual_log(logger, file_logger, 'info', f"   👀 Text preview: {preview}...")
                    else:
                        dual_log(logger, file_logger, 'warning', f"   ⚠️  No text extracted!")
                elif per_file_logging:
                    dual_log(logger, file_logger, 'info', f"📝 Text extracted: {len(extracted_text):,} chars via {extraction_method} ({text_time:.2f}s)")
                
                # Extract correspondence content (subject/body) - same as batch processor
                if log_text_extraction:
                    dual_log(logger, file_logger, 'info', f"📧 Starting correspondence extraction...")
                    
                from extract_correspondence_content import CorrespondenceExtractor
                extractor = CorrespondenceExtractor()
                
                corr_start = time.time()
                correspondence = extractor.extract_correspondence_content(extracted_text)
                corr_time = time.time() - corr_start
                
                # Create focused content like batch processor
                focused_content = f"Subject: {correspondence['subject']}\n\nContent: {correspondence['body']}"
                
                if log_text_extraction:
                    dual_log(logger, file_logger, 'info', f"✅ Correspondence extraction completed:")
                    dual_log(logger, file_logger, 'info', f"   📧 Subject length: {len(correspondence['subject'])} chars")
                    dual_log(logger, file_logger, 'info', f"   📄 Body length: {len(correspondence['body'])} chars")
                    dual_log(logger, file_logger, 'info', f"   📝 Focused content length: {len(focused_content)} chars")
                    dual_log(logger, file_logger, 'info', f"   ⏱️  Correspondence time: {corr_time:.2f}s")
                    if len(correspondence['subject']) > 0:
                        subject_preview = correspondence['subject'][:100].replace('\n', ' ')
                        dual_log(logger, file_logger, 'info', f"   📧 Subject preview: {subject_preview}...")
                    else:
                        dual_log(logger, file_logger, 'warning', f"   ⚠️  No subject extracted!")
                elif per_file_logging:
                    dual_log(logger, file_logger, 'info', f"📧 Correspondence: Subject({len(correspondence['subject'])}), Body({len(correspondence['body'])}) chars ({corr_time:.2f}s)")
                
                # Initialize result structure like batch processor
                file_result = {
                    'file_name': pdf_file.name,
                    'file_path': str(pdf_file),
                    'status': 'completed',
                    'processing_time': 0,
                    'approaches': {},
                    'subject': correspondence['subject'],
                    'body': correspondence['body'],
                    'text_info': {
                        'raw_length': len(extracted_text),
                        'focused_length': len(focused_content),
                        'extraction_method': extraction_method,
                        'correspondence_method': correspondence.get('extraction_method', 'pattern_matching')
                    },
                    'ground_truth': ground_truth_data.get(pdf_file.name, []),  # Use loaded ground truth data
                    'metrics': {}
                }
                
                # Process with hybrid_rag approach using the actual classifier
                approach_name = 'hybrid_rag'
                if classification_service.hybrid_rag_classifier:
                    if log_classification_details:
                        dual_log(logger, file_logger, 'info', f"🤖 Starting Hybrid RAG classification...")
                        dual_log(logger, file_logger, 'info', f"   📝 Input content length: {len(focused_content)} chars")
                        dual_log(logger, file_logger, 'info', f"   🔧 Classifier type: {type(classification_service.hybrid_rag_classifier).__name__}")
                    elif per_file_logging:
                        dual_log(logger, file_logger, 'info', f"🔍 Classifying with Hybrid RAG approach...")
                    else:
                        dual_log(logger, file_logger, 'info', f"  🔍 Classifying with Hybrid RAG approach...")
                    
                    approach_start = time.time()
                    approach_result = classification_service.hybrid_rag_classifier.classify(focused_content, is_file_path=False)
                    approach_time = time.time() - approach_start
                    
                    if log_classification_details:
                        logger.info(f"✅ Classification completed:")
                        logger.info(f"   ⏱️  Classification time: {approach_time:.2f}s")
                        logger.info(f"   📊 Result status: {approach_result.get('status', 'unknown')}")
                        logger.info(f"   🔬 Result keys: {list(approach_result.keys())}")
                        if 'categories' in approach_result:
                            logger.info(f"   📋 Raw categories count: {len(approach_result.get('categories', []))}")
                        if 'llm_provider_used' in approach_result:
                            logger.info(f"   🤖 LLM provider: {approach_result.get('llm_provider_used')}")
                        if 'method_used' in approach_result:
                            logger.info(f"   ⚙️  Method used: {approach_result.get('method_used')}")
                    
                    # Check if classification was skipped due to quality issues
                    if approach_result.get('status') == 'skipped':
                        skip_message = approach_result.get('message', 'Unknown reason')
                        if log_classification_details:
                            logger.warning(f"⚠️  Document skipped due to quality issues:")
                            logger.warning(f"   📝 Skip reason: {skip_message}")
                            logger.warning(f"   🔍 Quality check: {approach_result.get('quality_check', 'failed')}")
                            logger.warning(f"   ⏱️  Time spent: {approach_time:.2f}s")
                        elif per_file_logging:
                            logger.warning(f"⚠️  DOCUMENT SKIPPED: {skip_message} ({approach_time:.2f}s)")
                        else:
                            logger.warning(f"    ⚠️ Document skipped due to quality issues: {skip_message}")
                        
                        file_result['approaches'][approach_name] = {
                            'status': 'skipped',
                            'skip_reason': skip_message,
                            'quality_check': approach_result.get('quality_check', 'failed'),
                            'processing_time': approach_time,
                            'categories': [],  # No categories due to quality issues
                            'category_details': [],
                            'full_result': approach_result
                        }
                        
                        if not log_classification_details and not per_file_logging:
                            logger.info(f"    ⚠️ Hybrid RAG: {approach_time:.2f}s - DOCUMENT SKIPPED (Quality Issues)")
                    # Check if classification failed due to LLM error
                    elif approach_result.get('status') == 'error' and approach_result.get('error_type') == 'llm_validation_failed':
                        error_msg = f"❌ LLM Validation Failed ({approach_result.get('provider', 'Unknown')}): {approach_result.get('message', 'Unknown error')}"
                        logger.error(error_msg)
                        
                        file_result['approaches'][approach_name] = {
                            'status': 'llm_validation_failed',
                            'error_message': approach_result.get('message', 'LLM validation failed'),
                            'error_provider': approach_result.get('provider', 'unknown'),
                            'raw_semantic_results_count': approach_result.get('raw_semantic_results_count', 0),
                            'processing_time': approach_time,
                            'categories': [],  # No categories due to validation failure
                            'category_details': [],
                            'error_details': approach_result.get('error_details', {}),
                            'full_result': approach_result
                        }
                        
                        logger.info(f"    ❌ Hybrid RAG: {approach_time:.2f}s - LLM VALIDATION FAILED ({approach_result.get('provider', 'Unknown')})")
                    else:
                        # Extract categories with confidence scores, filtering by confidence threshold
                        categories = []
                        category_details = []
                        confidence_threshold = 0.2  # Same as batch processor
                        
                        raw_categories = approach_result.get('categories', [])
                        if log_classification_details:
                            logger.info(f"📋 Processing {len(raw_categories)} raw categories:")
                            for i, cat_info in enumerate(raw_categories[:5]):  # Show first 5
                                category = cat_info.get('category', '')
                                confidence = cat_info.get('confidence', 0.0)
                                logger.info(f"   {i+1}. {category} (confidence: {confidence:.3f})")
                            if len(raw_categories) > 5:
                                logger.info(f"   ... and {len(raw_categories) - 5} more")
                            logger.info(f"🔍 Applying confidence threshold: {confidence_threshold}")
                        
                        for cat_info in raw_categories:
                            category = cat_info.get('category', '')
                            confidence = cat_info.get('confidence', 0.0)
                            
                            # Only include categories with confidence >= threshold
                            if confidence >= confidence_threshold:
                                categories.append(category)
                                category_details.append({
                                    'category': category,
                                    'confidence': confidence,
                                    'evidence': cat_info.get('evidence', ''),  # RAG lookup evidence/justification
                                    'issue_types': cat_info.get('issue_types', [])  # Issue types that led to this category
                                })
                                if log_classification_details:
                                    logger.info(f"   ✅ Kept: {category} ({confidence:.3f})")
                            else:
                                if log_classification_details:
                                    logger.info(f"   ❌ Filtered: {category} ({confidence:.3f} < {confidence_threshold})")
                                else:
                                    logger.debug(f"Filtered out low confidence category '{category}' ({confidence:.3f} < {confidence_threshold})")
                        
                        if log_classification_details:
                            logger.info(f"📊 Category filtering results:")
                            logger.info(f"   📋 Raw categories: {len(raw_categories)}")
                            logger.info(f"   ✅ Kept categories: {len(categories)}")
                            logger.info(f"   📝 Final categories: {categories}")
                        elif per_file_logging:
                            logger.info(f"📊 Categories: {len(raw_categories)} → {len(categories)} (threshold: {confidence_threshold})")
                        else:
                            logger.info(f"    Kept {len(categories)} categories above {confidence_threshold} confidence threshold")
                        
                        file_result['approaches'][approach_name] = {
                            'status': 'success',
                            'categories': categories,
                            'category_details': category_details,  # Store detailed info with confidence
                            'processing_time': approach_time,
                            'provider_used': approach_result.get('llm_provider_used', approach_result.get('method_used', 'unknown')),
                            'full_result': approach_result
                        }
                        
                        if log_classification_details:
                            logger.info(f"✅ Hybrid RAG classification successful:")
                            logger.info(f"   ⏱️  Total time: {approach_time:.2f}s")
                            logger.info(f"   🤖 Provider: {approach_result.get('llm_provider_used', 'unknown')}")
                            logger.info(f"   📋 Final categories: {categories}")
                        elif per_file_logging:
                            logger.info(f"✅ Classification complete: {categories} ({approach_time:.2f}s)")
                        else:
                            logger.info(f"    ✅ Hybrid RAG: {approach_time:.2f}s - Categories: {categories}")
                else:
                    logger.warning("Hybrid RAG classifier not available")
                    file_result['approaches'][approach_name] = {
                        'status': 'error',
                        'error_message': 'Hybrid RAG classifier not initialized',
                        'categories': [],
                        'category_details': []
                    }
                
                # Calculate metrics if ground truth is available
                if file_result['ground_truth'] and enable_metrics:
                    gt_categories = file_result['ground_truth']
                    dual_log(logger, file_logger, 'info', f"📊 Ground truth loaded: {gt_categories}")
                    
                    # Calculate metrics for each approach using batch processor logic
                    for approach_name in file_result.get('approaches', {}):
                        if approach_name in file_result['approaches']:
                            predicted_categories = file_result['approaches'][approach_name].get('categories', [])
                            
                            # Filter out "Others" category for metrics calculation
                            gt_categories_filtered = [cat for cat in gt_categories if cat.lower() != 'others']
                            predicted_categories_filtered = [cat for cat in predicted_categories if cat.lower() != 'others']
                            
                            # Calculate metrics using MetricsCalculator
                            metrics = metrics_calculator.calculate_metrics(gt_categories_filtered, predicted_categories_filtered)
                            file_result['approaches'][approach_name]['metrics'] = metrics
                            
                            dual_log(logger, file_logger, 'info', f"📊 {approach_name} metrics calculated: precision={metrics.get('precision', 0):.3f}, recall={metrics.get('recall', 0):.3f}, f1={metrics.get('f1_score', 0):.3f}")
                    
                    dual_log(logger, file_logger, 'info', "📊 Metrics calculation completed")
                
                file_result['processing_time'] = time.time() - processing_start
                results['results'].append(file_result)
                results['processing_stats']['processed_files'] += 1
                
                # Final per-file summary
                if per_file_logging:
                    dual_log(logger, file_logger, 'info', f"\n📊 FILE SUMMARY: {pdf_file.name}")
                    dual_log(logger, file_logger, 'info', f"   ⏱️  Total processing time: {file_result['processing_time']:.2f}s")
                    dual_log(logger, file_logger, 'info', f"   📄 Status: {file_result['status']}")
                    if file_result.get('approaches'):
                        for approach_name, approach_data in file_result['approaches'].items():
                            status = approach_data.get('status', 'unknown')
                            categories = approach_data.get('categories', [])
                            proc_time = approach_data.get('processing_time', 0)
                            dual_log(logger, file_logger, 'info', f"   🔍 {approach_name}: {status} - {len(categories)} categories ({proc_time:.2f}s)")
                            if categories and log_classification_details:
                                dual_log(logger, file_logger, 'info', f"      📋 Categories: {categories}")
                    if file_result.get('ground_truth'):
                        dual_log(logger, file_logger, 'info', f"   📊 Ground truth: {file_result['ground_truth']}")
                    dual_log(logger, file_logger, 'info', f"{'='*80}\n")
                else:
                    dual_log(logger, file_logger, 'info', f"✅ {pdf_file.name} processed successfully")
                
                # Final log message for individual file log
                if file_logger:
                    file_logger.info(f"🎯 PROCESSING COMPLETED: {pdf_file.name}")
                    file_logger.info(f"📊 Final Status: {file_result.get('status', 'unknown')}")
                    file_logger.info(f"⏱️  Total Time: {file_result.get('processing_time', 0):.2f}s")
                    
            except Exception as e:
                dual_log(logger, file_logger, 'error', f"❌ Failed to process {pdf_file.name}: {e}")
                if file_logger:
                    file_logger.error(f"🎯 PROCESSING FAILED: {pdf_file.name}")
                    file_logger.error(f"❌ Error: {str(e)}")
                file_result = {
                    'file_name': pdf_file.name,
                    'file_path': str(pdf_file),
                    'status': 'failed',
                    'error': str(e)
                }
                results['results'].append(file_result)
                results['processing_stats']['failed_files'] += 1
                
            finally:
                # Clean up per-file logger
                if file_logger:
                    cleanup_per_file_logger(file_logger)
        
        results['processing_stats']['end_time'] = datetime.now()
        
        # Add config and overall metrics information like batch processor
        results['config'] = {
            'batch_config': {
                'batch_processing': {
                    'approaches': {
                        'hybrid_rag': {
                            'enabled': True,
                            'priority': 1
                        },
                        'pure_llm': {
                            'enabled': False,  # Currently not enabled in integrated backend
                            'priority': 2
                        }
                    },
                    'enabled': True,
                    'evaluation': {
                        'auto_detect_ground_truth': False,
                        'enabled': enable_metrics,
                        'ground_truth_patterns': [
                            'EDMS*.xlsx',
                            'LOT-*.xlsx',
                            'ground_truth*.xlsx',
                            '*_labels.xlsx'
                        ]
                    },
                    'output': {
                        'results_folder': output_folder,
                        'save_format': 'xlsx'
                    },
                    'processing': {
                        'max_files': max_files,
                        'max_pages_per_pdf': 2,
                        'rate_limit_delay': 1,
                        'skip_on_error': True
                    }
                }
            },
            'enabled_approaches': ['hybrid_rag']
        }
        
        # Generate Excel and JSON files
        excel_files = []
        json_files = []
        
        try:
            # Create output directory
            output_path = Path(output_folder)
            output_path.mkdir(parents=True, exist_ok=True)
            
            # Generate timestamp for unique filenames
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Save to Excel
            excel_file = output_path / f"batch_results_{timestamp}.xlsx"
            _save_batch_results_to_excel(results, excel_file, ground_truth_file)
            excel_files.append(str(excel_file))
            logger.info(f"💾 Results saved to Excel: {excel_file}")
            
            # Save to JSON for programmatic access
            json_file = output_path / f"batch_results_{timestamp}.json"
            with open(json_file, 'w', encoding='utf-8') as f:
                import json
                json.dump(results, f, indent=2, default=str)
            json_files.append(str(json_file))
            logger.info(f"💾 Results saved to JSON: {json_file}")
            
        except Exception as save_error:
            logger.error(f"Failed to save results files: {save_error}")
            excel_files = [f"Error saving Excel: {save_error}"]
            json_files = [f"Error saving JSON: {save_error}"]
        
        return jsonify({
            "status": "success",
            "success": True,
            "data": {
                "processing_stats": results.get('processing_stats', {}),
                "total_files": len(pdf_files),
                "results": results.get('results', []),
                "overall_metrics": results.get('overall_metrics', {}),
                "output_folder": output_folder,
                "excel_files": excel_files,
                "json_files": json_files
            },
            "message": f"Successfully processed {len(pdf_files)} PDF files and saved results to {output_folder}"
        })
        
    except Exception as e:
        logger.error(f"Batch folder processing failed: {e}")
        return jsonify({"error": str(e)}), 500

@service_api.route('/hybrid-rag-classification/process-single-pdf', methods=['POST'])
def process_single_pdf():
    """Process a single PDF file with batch-like processing and file output"""
    try:
        data = request.get_json()
        if not data or 'pdf_path' not in data:
            return jsonify({"error": "pdf_path is required"}), 400
        
        pdf_path = data['pdf_path']
        
        # Validate file exists
        if not Path(pdf_path).exists():
            return jsonify({"error": f"PDF file not found: {pdf_path}"}), 404
        
        # Get options
        approaches = data.get('approaches', ['hybrid_rag'])
        confidence_threshold = data.get('confidence_threshold', 0.3)
        max_pages = data.get('max_pages', 2)
        ground_truth_file = data.get('ground_truth_file')
        output_folder = data.get('output_folder', './results/api_processing')
        save_files = data.get('save_files', True)
        
        # Use the unified processor
        from unified_pdf_processor import UnifiedPDFProcessor
        
        # Initialize processor
        processor = UnifiedPDFProcessor()
        
        # Process single PDF
        result = processor.process_single_pdf(
            pdf_path=pdf_path,
            approaches=approaches,
            confidence_threshold=confidence_threshold,
            max_pages=max_pages,
            ground_truth_file=ground_truth_file
        )
        
        # Save files if requested
        output_files = []
        if save_files:
            from unified_single_pdf_processor import save_results_to_json, save_results_to_excel
            
            json_file = save_results_to_json(result, output_folder)
            excel_file = save_results_to_excel(result, output_folder)
            
            if json_file:
                output_files.append(json_file)
            if excel_file:
                output_files.append(excel_file)
        
        return jsonify({
            "status": "success",
            "success": True,
            "data": {
                "result": result,
                "output_files": output_files,
                "output_folder": output_folder
            },
            "message": f"Successfully processed {Path(pdf_path).name}"
        })
        
    except Exception as e:
        logger.error(f"Single PDF processing failed: {e}")
        return jsonify({"error": str(e)}), 500


def _save_batch_results_to_excel(results: Dict, excel_path: Path, ground_truth_file: str = None):
    """Save batch results to Excel file with multiple sheets, similar to batch_processor format"""
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Summary sheet (Results)
            summary_data = []
            for result in results['results']:
                if result['status'] == 'completed':
                    row = {
                        'File Name': result['file_name'],
                        'Subject': result.get('subject', ''),
                        'Body': result.get('body', ''),
                        'Processing Time (s)': f"{result['processing_time']:.2f}",
                        'Ground Truth': ', '.join(result.get('ground_truth', [])),
                    }
                    
                    # Add approach results - same as batch processor format
                    for approach_name, approach_data in result.get('approaches', {}).items():
                        approach_title = approach_name.replace("_", " ").title()
                        
                        # Check if this approach failed due to LLM validation error
                        if approach_data.get('status') == 'llm_validation_failed':
                            # Handle LLM validation failure in summary
                            provider = approach_data.get('error_provider', 'Unknown')
                            raw_results = approach_data.get('raw_semantic_results_count', 0)
                            
                            row['Predicted Categories'] = f'ERROR: LLM Validation Failed ({provider})'
                            row['Categories with Confidence'] = f'Found {raw_results} semantic results but LLM validation failed'
                            row['RAG Time (s)'] = f"{approach_data['processing_time']:.2f}"
                        else:
                            # Normal processing - Categories and confidence scores
                            categories_with_confidence = []
                            for cat_detail in approach_data.get('category_details', []):
                                cat = cat_detail.get('category', '')
                                conf = cat_detail.get('confidence', 0.0)
                                
                                # Try to get original LLM confidence if available
                                original_conf = None
                                if 'full_result' in approach_data and 'identified_issues' in approach_data['full_result']:
                                    for issue in approach_data['full_result']['identified_issues']:
                                        if issue.get('source') == 'llm_validation':
                                            original_conf = issue.get('original_confidence', None)
                                            break
                                
                                # Format confidence display
                                if original_conf and original_conf != conf:
                                    categories_with_confidence.append(f"{cat} ({conf:.3f}, LLM: {original_conf:.3f})")
                                else:
                                    categories_with_confidence.append(f"{cat} ({conf:.3f})")
                            
                            row['Predicted Categories'] = ', '.join(approach_data['categories'])
                            row['Categories with Confidence'] = ', '.join(categories_with_confidence)
                            row['RAG Time (s)'] = f"{approach_data['processing_time']:.2f}"
                        
                        # Add metrics if available
                        if 'metrics' in approach_data:
                            metrics = approach_data['metrics']
                            row['Precision'] = round(metrics.get('precision', 0), 2) if metrics.get('precision') is not None else None
                            row['Recall'] = round(metrics.get('recall', 0), 2) if metrics.get('recall') is not None else None
                            row['F1'] = round(metrics.get('f1_score', 0), 2) if metrics.get('f1_score') is not None else None
                            row['Exact Match'] = round(metrics.get('exact_match', 0), 2) if metrics.get('exact_match') is not None else None
                            row['True Positives'] = int(metrics.get('tp', 0)) if metrics.get('tp') is not None else None
                            row['False Positives'] = int(metrics.get('fp', 0)) if metrics.get('fp') is not None else None
                            row['False Negatives'] = int(metrics.get('fn', 0)) if metrics.get('fn') is not None else None
                            row['False Negatives List'] = ', '.join(metrics.get('missed_categories', [])) if metrics.get('missed_categories') else ''
                            row['Jaccard Similarity'] = round(metrics.get('jaccard_similarity', 0), 2) if metrics.get('jaccard_similarity') is not None else None
                    
                    summary_data.append(row)
                else:
                    summary_data.append({
                        'File Name': result['file_name'],
                        'Status': result['status'],
                        'Error': result.get('error', '')
                    })
            
            summary_df = pd.DataFrame(summary_data)
            summary_df = summary_df.sort_values('File Name') if not summary_df.empty else summary_df
            summary_df.to_excel(writer, sheet_name='Results', index=False)
            
            # Format the Results sheet with auto-filter and left-alignment
            if 'Results' in writer.sheets:
                worksheet = writer.sheets['Results']
                if worksheet.max_row > 1:  # Only if there's data
                    worksheet.auto_filter.ref = worksheet.dimensions
                
                # Apply left alignment to all cells
                from openpyxl.styles import Alignment
                left_align = Alignment(horizontal='left', vertical='top', wrap_text=False)
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = left_align
            
            # Detailed results sheet with one row per category prediction - same as batch processor
            detailed_data = []
            for result in results['results']:
                if result['status'] == 'completed':
                    base_info = {
                        'File Name': result['file_name'],
                        'Subject': result.get('subject', ''),
                        'Body': result.get('body', '')[:500] + '...' if len(result.get('body', '')) > 500 else result.get('body', ''),
                        'Ground Truth': ', '.join(result.get('ground_truth', []))
                    }
                    
                    # Add one row per approach per category - same as batch processor format
                    for approach_name, approach_data in result.get('approaches', {}).items():
                        approach_title = "RAG"  # Simplified approach name
                        
                        # Check if this approach failed due to LLM validation error
                        if approach_data.get('status') == 'llm_validation_failed':
                            # Handle LLM validation failure
                            detailed_row = base_info.copy()
                            error_msg = approach_data.get('error_message', 'LLM validation failed')
                            provider = approach_data.get('error_provider', 'Unknown')
                            raw_results = approach_data.get('raw_semantic_results_count', 0)
                            
                            detailed_row.update({
                                'Approach': approach_title,
                                'Predicted Category': 'ERROR: LLM Validation Failed',
                                'Confidence Score': 0.0,
                                'Issue Types': f'LLM Error ({provider})',
                                'Justification': f'LLM validation failed for {provider}. Found {raw_results} semantic results but could not validate. Error: {error_msg[:200]}...' if len(error_msg) > 200 else f'LLM validation failed for {provider}. Found {raw_results} semantic results but could not validate. Error: {error_msg}',
                                'Processing Time (s)': f"{approach_data['processing_time']:.2f}"
                            })
                            detailed_data.append(detailed_row)
                            
                        elif approach_data.get('category_details'):
                            for cat_detail in approach_data['category_details']:
                                detailed_row = base_info.copy()
                                
                                # Get evidence from RAG lookup as justification, truncate if too long
                                evidence = cat_detail.get('evidence', 'No supporting evidence found in RAG lookup')
                                # Truncate long evidence and show only the first sentence or first 150 characters
                                if evidence and len(evidence) > 150:
                                    sentences = evidence.split('. ')
                                    justification = sentences[0] + ('.' if len(sentences[0]) < len(evidence) else '') + '...'
                                else:
                                    justification = evidence
                                
                                # Get issue types that led to this category
                                issue_types = cat_detail.get('issue_types', [])
                                issue_types_str = ', '.join(issue_types) if issue_types else 'No issue types found'
                                
                                # Try to get original LLM confidence if available
                                original_conf = None
                                llm_confidence_note = ''
                                if 'full_result' in approach_data and 'identified_issues' in approach_data['full_result']:
                                    for issue in approach_data['full_result']['identified_issues']:
                                        if issue.get('source') == 'llm_validation':
                                            original_conf = issue.get('original_confidence', None)
                                            if original_conf and original_conf != cat_detail.get('confidence', 0.0):
                                                llm_confidence_note = f" (Original LLM: {original_conf:.3f})"
                                            break
                                
                                detailed_row.update({
                                    'Approach': approach_title,
                                    'Predicted Category': cat_detail.get('category', ''),
                                    'Confidence Score': f"{cat_detail.get('confidence', 0.0):.3f}{llm_confidence_note}",
                                    'Issue Types': issue_types_str,
                                    'Justification': justification,
                                    'Processing Time (s)': f"{approach_data['processing_time']:.2f}"
                                })
                                detailed_data.append(detailed_row)
                        else:
                            # No categories found (but no error)
                            detailed_row = base_info.copy()
                            detailed_row.update({
                                'Approach': approach_title,
                                'Predicted Category': 'No categories found',
                                'Confidence Score': 0.0,
                                'Issue Types': 'N/A',
                                'Justification': 'No categories were predicted above the confidence threshold',
                                'Processing Time (s)': f"{approach_data['processing_time']:.2f}"
                            })
                            detailed_data.append(detailed_row)
            
            if detailed_data:
                detailed_df = pd.DataFrame(detailed_data)
                detailed_df = detailed_df.sort_values('File Name') if not detailed_df.empty else detailed_df
                detailed_df.to_excel(writer, sheet_name='Detailed Results', index=False)
                
                # Format the Detailed Results sheet
                if 'Detailed Results' in writer.sheets:
                    detailed_worksheet = writer.sheets['Detailed Results']
                    if detailed_worksheet.max_row > 1:
                        detailed_worksheet.auto_filter.ref = detailed_worksheet.dimensions
                    
                    from openpyxl.styles import Alignment
                    left_align = Alignment(horizontal='left', vertical='top', wrap_text=False)
                    for row in detailed_worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = left_align
            
            # Processing stats sheet
            stats_data = []
            processing_stats = results.get('processing_stats', {})
            for key, value in processing_stats.items():
                stats_data.append({
                    'Metric': key.replace('_', ' ').title(),
                    'Value': str(value)
                })
            
            if stats_data:
                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, sheet_name='Processing Stats', index=False)
                
                # Format the Processing Stats sheet
                if 'Processing Stats' in writer.sheets:
                    stats_worksheet = writer.sheets['Processing Stats']
                    from openpyxl.styles import Alignment
                    left_align = Alignment(horizontal='left', vertical='top', wrap_text=False)
                    for row in stats_worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = left_align
        
        # Chunk Data Review sheet for debugging - same as batch processor
        chunk_debug_data = []
        for result in results['results']:
            if result['status'] == 'completed':
                for approach_name, approach_data in result.get('approaches', {}).items():
                    # Extract chunk debug data if available
                    if 'full_result' in approach_data and 'chunk_debug_data' in approach_data['full_result']:
                        chunk_data = approach_data['full_result']['chunk_debug_data']
                        for chunk in chunk_data:
                            chunk_debug_data.append({
                                'File Name': result['file_name'],
                                'Approach': approach_name.replace('_', ' ').title(),
                                'Chunk ID': chunk.get('chunk_id', ''),
                                'Chunk Length (chars)': chunk.get('chunk_length', 0),
                                'Start Position': chunk.get('start_pos', 0),
                                'End Position': chunk.get('end_pos', 0),
                                'Chunk Text Preview': chunk.get('chunk_text', ''),
                                'Search Results Count': chunk.get('search_results_count', 0),
                                'Unique Issues Found': chunk.get('unique_issues_found', 0),
                                'Issues List': chunk.get('issues_list', ''),
                                'Top 3 Similarities': ', '.join([f'{sim:.3f}' for sim in chunk.get('top_similarities', [])]),
                                'Average Similarity': f"{chunk.get('avg_similarity', 0):.3f}"
                            })
        
        if chunk_debug_data:
            chunk_df = pd.DataFrame(chunk_debug_data)
            # Sort by File Name, then Chunk ID
            chunk_df = chunk_df.sort_values(['File Name', 'Chunk ID']) if not chunk_df.empty else chunk_df
            chunk_df.to_excel(writer, sheet_name='Chunk_Data_Review', index=False)
            
            # Format the Chunk Data Review sheet
            if 'Chunk_Data_Review' in writer.sheets:
                chunk_worksheet = writer.sheets['Chunk_Data_Review']
                if chunk_worksheet.max_row > 1:
                    chunk_worksheet.auto_filter.ref = chunk_worksheet.dimensions
                
                from openpyxl.styles import Alignment
                left_align = Alignment(horizontal='left', vertical='top', wrap_text=False)
                for row in chunk_worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = left_align
            
            logger.info(f"📊 Chunk debug data saved: {len(chunk_debug_data)} chunk entries across {len(set(d['File Name'] for d in chunk_debug_data))} files")
        else:
            logger.info("ℹ️ No chunk debug data available (only available with hybrid_rag approach)")
        
        logger.info(f"📊 Excel file created with {len(summary_data)} results")
        
    except Exception as e:
        logger.error(f"Failed to create Excel file: {e}")
        raise