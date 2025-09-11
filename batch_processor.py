#!/usr/bin/env python3
"""
Batch PDF Processing System
Configurable batch processing for contract correspondence classification
"""

import os
import time
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import json
from datetime import datetime
import yaml

from classifier.config_manager import ConfigManager
from classifier.issue_mapper import IssueCategoryMapper
from classifier.validation import ValidationEngine
from classifier.data_sufficiency import DataSufficiencyAnalyzer
from classifier.pure_llm import PureLLMClassifier
from classifier.hybrid_rag import HybridRAGClassifier
from classifier.pdf_extractor import PDFExtractor
from classifier.category_normalizer import CategoryNormalizer
from extract_correspondence_content import CorrespondenceExtractor
from metrics_calculator import MetricsCalculator

logger = logging.getLogger(__name__)


class BatchPDFProcessor:
    """
    Configurable batch PDF processing system for contract correspondence classification.
    
    Features:
    - Configurable approach enable/disable (Pure LLM, Hybrid RAG)
    - Optional ground truth evaluation with auto-detection
    - Flexible input/output handling
    - Performance metrics and reporting
    - Error handling and recovery
    """
    
    def __init__(self, config_path: str = None, batch_config_path: str = None):
        """
        Initialize the batch processor.
        
        Args:
            config_path: Path to main configuration file (config.yaml)
            batch_config_path: Path to batch configuration file (batch_config.yaml)
        """
        # Load configurations
        self.config_manager = ConfigManager(config_path or "config.yaml")
        self.config = self.config_manager.get_all_config()
        
        # Load batch-specific configuration
        self.batch_config = self._load_batch_config(batch_config_path or "batch_config.yaml")
        
        # Initialize components
        self._init_components()
        
        # Initialize classifiers based on configuration
        self.classifiers = {}
        self._init_classifiers()
        
        # Processing state
        self.results = []
        self.processing_stats = {
            'total_files': 0,
            'processed_files': 0,
            'failed_files': 0,
            'start_time': None,
            'end_time': None
        }
    
    def _load_batch_config(self, config_path: str) -> Dict:
        """Load batch configuration from YAML file."""
        try:
            config_path = Path(config_path)
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    return yaml.safe_load(f)
            else:
                logger.warning(f"Batch config file not found: {config_path}, using defaults")
                return self._get_default_batch_config()
        except Exception as e:
            logger.error(f"Error loading batch config: {e}")
            return self._get_default_batch_config()
    
    def _get_default_batch_config(self) -> Dict:
        """Get default batch configuration."""
        return {
            'batch_processing': {
                'enabled': True,
                'approaches': {
                    'hybrid_rag': {'enabled': True, 'priority': 1},
                    'pure_llm': {'enabled': False, 'priority': 2}
                },
                'evaluation': {
                    'enabled': True,
                    'auto_detect_ground_truth': True,
                    'ground_truth_patterns': ["EDMS*.xlsx", "ground_truth*.xlsx", "*_labels.xlsx"]
                },
                'output': {
                    'results_folder': 'results',
                    'save_format': 'xlsx'
                },
                'processing': {
                    'max_pages_per_pdf': 2,
                    'skip_on_error': True,
                    'rate_limit_delay': 3
                }
            }
        }
    
    def _init_components(self):
        """Initialize shared components."""
        training_data_path = Path(self.config['data']['training_data'])
        
        self.issue_mapper = IssueCategoryMapper(training_data_path)
        self.validator = ValidationEngine(training_data_path)
        self.data_analyzer = DataSufficiencyAnalyzer(training_data_path)
        self.pdf_extractor = PDFExtractor(
            max_pages=self.batch_config['batch_processing']['processing'].get('max_pages_per_pdf', 2)
        )
        self.correspondence_extractor = CorrespondenceExtractor()
        self.category_normalizer = CategoryNormalizer(strict_mode=False)
        
        # Initialize metrics calculator if evaluation is enabled
        if self.batch_config['batch_processing']['evaluation']['enabled']:
            self.metrics_calculator = MetricsCalculator()
        else:
            self.metrics_calculator = None
        
        logger.info("Batch processor components initialized")
    
    def _init_classifiers(self):
        """Initialize classifiers based on configuration."""
        batch_approaches = self.batch_config['batch_processing']['approaches']
        
        # Initialize Hybrid RAG if enabled
        if batch_approaches.get('hybrid_rag', {}).get('enabled', False):
            if self.config['approaches']['hybrid_rag']['enabled']:
                self.classifiers['hybrid_rag'] = HybridRAGClassifier(
                    config=self.config['approaches']['hybrid_rag'],
                    issue_mapper=self.issue_mapper,
                    validator=self.validator,
                    data_analyzer=self.data_analyzer
                )
                logger.info("✅ Hybrid RAG classifier initialized")
            else:
                logger.warning("⚠️ Hybrid RAG requested but disabled in main config")
        
        # Initialize Pure LLM if enabled
        if batch_approaches.get('pure_llm', {}).get('enabled', False):
            if self.config['approaches']['pure_llm']['enabled']:
                self.classifiers['pure_llm'] = PureLLMClassifier(
                    config=self.config['approaches']['pure_llm'],
                    issue_mapper=self.issue_mapper,
                    validator=self.validator,
                    data_analyzer=self.data_analyzer
                )
                logger.info("✅ Pure LLM classifier initialized")
            else:
                logger.warning("⚠️ Pure LLM requested but disabled in main config")
        
        if not self.classifiers:
            raise ValueError("No classifiers enabled! Please enable at least one approach.")
        
        logger.info(f"🚀 Initialized {len(self.classifiers)} classifiers: {list(self.classifiers.keys())}")
    
    def process_pdf_folder(self, 
                          pdf_folder: str, 
                          ground_truth_file: str = None,
                          output_folder: str = None) -> Dict:
        """
        Process all PDFs in a folder with optional ground truth evaluation.
        
        Args:
            pdf_folder: Path to folder containing PDF files
            ground_truth_file: Optional path to ground truth Excel file
            output_folder: Optional custom output folder
            
        Returns:
            Dictionary containing processing results and metrics
        """
        pdf_folder = Path(pdf_folder)
        if not pdf_folder.exists():
            raise FileNotFoundError(f"PDF folder not found: {pdf_folder}")
        
        # Auto-detect ground truth if enabled and not provided
        if ground_truth_file is None and self.batch_config['batch_processing']['evaluation']['auto_detect_ground_truth']:
            ground_truth_file = self._auto_detect_ground_truth(pdf_folder)
        
        # Load ground truth if available
        ground_truth = None
        if ground_truth_file and self.metrics_calculator:
            ground_truth = self._load_ground_truth(ground_truth_file)
            logger.info(f"📊 Loaded ground truth for {len(ground_truth)} files")
        elif self.metrics_calculator:
            logger.info("📊 Metrics calculation enabled but no ground truth found")
        
        # Find PDF files
        pdf_files = list(pdf_folder.glob("*.pdf"))
        if not pdf_files:
            raise ValueError(f"No PDF files found in {pdf_folder}")
        
        logger.info(f"📄 Found {len(pdf_files)} PDF files to process")
        
        # Initialize processing
        self.processing_stats['total_files'] = len(pdf_files)
        self.processing_stats['start_time'] = datetime.now()
        self.results = []
        
        # Process each PDF
        for i, pdf_file in enumerate(pdf_files, 1):
            logger.info(f"📄 [{i}/{len(pdf_files)}] Processing: {pdf_file.name}")
            
            try:
                result = self._process_single_pdf(pdf_file, ground_truth)
                self.results.append(result)
                self.processing_stats['processed_files'] += 1
                
                # Rate limiting
                rate_limit = self.batch_config['batch_processing']['processing'].get('rate_limit_delay', 3)
                if i < len(pdf_files) and rate_limit > 0:  # Don't delay after last file
                    logger.debug(f"⏳ Rate limiting delay ({rate_limit}s)...")
                    time.sleep(rate_limit)
                    
            except Exception as e:
                logger.error(f"❌ Failed to process {pdf_file.name}: {e}")
                self.processing_stats['failed_files'] += 1
                
                if not self.batch_config['batch_processing']['processing'].get('skip_on_error', True):
                    raise
                    
                # Add failed result placeholder
                self.results.append({
                    'file_name': pdf_file.name,
                    'status': 'failed',
                    'error': str(e),
                    'approaches': {}
                })
        
        self.processing_stats['end_time'] = datetime.now()
        
        # Generate final results
        final_results = self._generate_final_results(ground_truth, output_folder)
        
        logger.info(f"🎉 Batch processing completed!")
        logger.info(f"📊 Processed: {self.processing_stats['processed_files']}/{self.processing_stats['total_files']} files")
        logger.info(f"⏱️  Total time: {self.processing_stats['end_time'] - self.processing_stats['start_time']}")
        
        return final_results
    
    def _auto_detect_ground_truth(self, pdf_folder: Path) -> Optional[str]:
        """Auto-detect ground truth files in the PDF folder or parent directories."""
        patterns = self.batch_config['batch_processing']['evaluation']['ground_truth_patterns']
        
        # Search in PDF folder and parent directories
        search_paths = [pdf_folder, pdf_folder.parent, pdf_folder.parent.parent]
        
        for search_path in search_paths:
            for pattern in patterns:
                matches = list(search_path.glob(pattern))
                if matches:
                    ground_truth_file = matches[0]
                    logger.info(f"📊 Auto-detected ground truth: {ground_truth_file}")
                    return str(ground_truth_file)
        
        logger.info("📊 No ground truth file auto-detected")
        return None
    
    def _load_ground_truth(self, ground_truth_file: str) -> Dict:
        """Load ground truth from Excel file."""
        try:
            df = pd.read_excel(ground_truth_file)
            
            # Auto-detect the format and find the correct columns
            ground_truth = {}
            
            # Check if this looks like the LOT-21 format (has "Sr. No" in first data row)
            if len(df) > 1 and str(df.iloc[1, 0]).strip() == "Sr. No":
                logger.info("📊 Detected LOT-21 ground truth format")
                # LOT-21 format: skip first 2 rows (headers), file name in column 2, categories in column 5
                
                for idx in range(2, len(df)):
                    row = df.iloc[idx]
                    
                    # File name is in column 2, add .pdf extension if missing
                    file_name_raw = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ""
                    if not file_name_raw or file_name_raw.strip() in ["", "nan"]:
                        continue
                    
                    file_name = file_name_raw.strip()
                    if not file_name.lower().endswith('.pdf'):
                        file_name = file_name + '.pdf'
                    
                    # Categories are in column 5, comma-separated
                    categories_raw = str(row.iloc[5]) if pd.notna(row.iloc[5]) else ""
                    
                    if categories_raw and categories_raw.strip() not in ["", "nan"]:
                        # Use the CategoryNormalizer to parse and normalize categories
                        normalized_categories = self.category_normalizer.parse_and_normalize_categories(categories_raw)
                        
                        # Consolidate categories for the same file
                        if file_name in ground_truth:
                            # Add new categories to existing ones, avoiding duplicates
                            for cat in normalized_categories:
                                if cat not in ground_truth[file_name]:
                                    ground_truth[file_name].append(cat)
                        else:
                            # First occurrence of this file
                            ground_truth[file_name] = normalized_categories
            
            else:
                # Generic format: first column is file name, remaining columns are categories  
                logger.info("📊 Using generic ground truth format")
                for _, row in df.iterrows():
                    file_name = str(row.iloc[0])  # First column is file name
                    if not file_name or file_name.strip() in ["", "nan"]:
                        continue
                        
                    raw_categories = []
                    # Collect non-empty categories from remaining columns
                    for col_idx in range(1, len(row)):
                        if pd.notna(row.iloc[col_idx]) and str(row.iloc[col_idx]).strip():
                            cat = str(row.iloc[col_idx]).strip()
                            if cat not in ["", "nan"]:
                                raw_categories.append(cat)
                    
                    # Normalize categories using CategoryNormalizer
                    normalized_categories = []
                    for cat in raw_categories:
                        norm_cat, status, confidence = self.category_normalizer.normalize_category(cat)
                        if norm_cat and norm_cat not in normalized_categories:
                            normalized_categories.append(norm_cat)
                    
                    # Consolidate categories for the same file
                    if file_name in ground_truth:
                        # Add new categories to existing ones, avoiding duplicates
                        for cat in normalized_categories:
                            if cat not in ground_truth[file_name]:
                                ground_truth[file_name].append(cat)
                    else:
                        # First occurrence of this file
                        ground_truth[file_name] = normalized_categories
            
            # Remove files with no categories
            ground_truth = {k: v for k, v in ground_truth.items() if v}
            
            logger.info(f"📊 Loaded ground truth for {len(ground_truth)} files")
            if logger.isEnabledFor(logging.DEBUG):
                for file_name, categories in list(ground_truth.items())[:5]:  # Show first 5
                    logger.debug(f"📊 {file_name}: {categories}")
            
            return ground_truth
            
        except Exception as e:
            logger.error(f"Failed to load ground truth from {ground_truth_file}: {e}")
            return {}
    
    def _process_single_pdf(self, pdf_file: Path, ground_truth: Dict = None) -> Dict:
        """Process a single PDF file with all enabled approaches."""
        result = {
            'file_name': pdf_file.name,
            'file_path': str(pdf_file),
            'status': 'completed',
            'processing_time': 0,
            'approaches': {},
            'ground_truth': ground_truth.get(pdf_file.name, []) if ground_truth else [],
            'metrics': {}
        }
        
        start_time = time.time()
        
        try:
            # Extract text and correspondence content
            raw_text, extraction_method = self.pdf_extractor.extract_text(pdf_file)
            extraction_result = self.correspondence_extractor.extract_correspondence_content(raw_text)
            
            focused_content = f"Subject: {extraction_result['subject']}\n\nContent: {extraction_result['body']}"
            
            result['text_info'] = {
                'raw_length': len(raw_text),
                'focused_length': len(focused_content),
                'extraction_method': extraction_method,
                'correspondence_method': extraction_result['extraction_method']
            }
            
            # Store extracted subject and body
            result['subject'] = extraction_result['subject']
            result['body'] = extraction_result['body']
            
            # Process with each enabled approach
            for approach_name, classifier in self.classifiers.items():
                logger.info(f"  🔍 Classifying with {approach_name.replace('_', ' ').title()} approach...")
                
                approach_start = time.time()
                approach_result = classifier.classify(focused_content, is_file_path=False)
                approach_time = time.time() - approach_start
                
                # Extract categories with confidence scores, filtering by confidence threshold
                categories = []
                category_details = []
                confidence_threshold = 0.5
                
                for cat_info in approach_result.get('categories', []):
                    category = cat_info.get('category', '')
                    confidence = cat_info.get('confidence', 0.0)
                    
                    # Only include categories with confidence >= threshold
                    if confidence >= confidence_threshold:
                        categories.append(category)
                        category_details.append({
                            'category': category,
                            'confidence': confidence,
                            'evidence': cat_info.get('evidence', ''),  # RAG lookup evidence/justification
                            'issue_types': cat_info.get('issue_types', [])  # Issue types that led to this category
                        })
                    else:
                        logger.debug(f"Filtered out low confidence category '{category}' ({confidence:.3f} < {confidence_threshold})")
                
                logger.info(f"    Kept {len(categories)} categories above {confidence_threshold} confidence threshold")
                
                result['approaches'][approach_name] = {
                    'categories': categories,
                    'category_details': category_details,  # Store detailed info with confidence
                    'processing_time': approach_time,
                    'provider_used': approach_result.get('llm_provider_used', approach_result.get('method_used', 'unknown')),
                    'full_result': approach_result
                }
                
                logger.info(f"    ✅ {approach_name.replace('_', ' ').title()}: {approach_time:.2f}s - Categories: {categories}")
            
            # Calculate metrics if ground truth is available and metrics calculator is enabled
            if self.metrics_calculator and ground_truth and pdf_file.name in ground_truth:
                gt_categories = ground_truth[pdf_file.name]
                
                for approach_name in result['approaches']:
                    predicted_categories = result['approaches'][approach_name]['categories']
                    
                    # Filter out "Others" category for metrics calculation
                    gt_categories_filtered = [cat for cat in gt_categories if cat.lower() != 'others']
                    predicted_categories_filtered = [cat for cat in predicted_categories if cat.lower() != 'others']
                    
                    metrics = self.metrics_calculator.calculate_metrics(gt_categories_filtered, predicted_categories_filtered)
                    result['approaches'][approach_name]['metrics'] = metrics
            
            result['processing_time'] = time.time() - start_time
            
        except Exception as e:
            result['status'] = 'failed'
            result['error'] = str(e)
            result['processing_time'] = time.time() - start_time
            raise
        
        return result
    
    def _generate_final_results(self, ground_truth: Dict = None, output_folder: str = None) -> Dict:
        """Generate and save final results."""
        if not output_folder:
            output_folder = self.batch_config['batch_processing']['output']['results_folder']
        
        output_path = Path(output_folder)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # Generate timestamp for unique filenames
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Prepare results for saving
        final_results = {
            'processing_stats': self.processing_stats,
            'config': {
                'batch_config': self.batch_config,
                'enabled_approaches': list(self.classifiers.keys())
            },
            'results': self.results
        }
        
        # Calculate overall metrics if ground truth is available
        if self.metrics_calculator and ground_truth:
            overall_metrics = self._calculate_overall_metrics(ground_truth)
            final_results['overall_metrics'] = overall_metrics
        
        # Save results in requested format
        save_format = self.batch_config['batch_processing']['output'].get('save_format', 'xlsx')
        
        if save_format == 'xlsx':
            excel_path = output_path / f"batch_results_{timestamp}.xlsx"
            self._save_results_excel(final_results, excel_path)
            logger.info(f"💾 Results saved to Excel: {excel_path}")
        
        # Always save JSON for programmatic access
        json_path = output_path / f"batch_results_{timestamp}.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(final_results, f, indent=2, default=str)
        logger.info(f"💾 Results saved to JSON: {json_path}")
        
        return final_results
    
    def _calculate_overall_metrics(self, ground_truth: Dict) -> Dict:
        """Calculate overall metrics across all files and approaches, excluding 'Others' category."""
        overall_metrics = {}
        
        for approach_name in self.classifiers.keys():
            all_gt = []
            all_pred = []
            
            for result in self.results:
                if result['status'] == 'completed' and approach_name in result['approaches']:
                    file_name = result['file_name']
                    if file_name in ground_truth:
                        # Filter out "Others" category for metrics calculation
                        gt_categories_filtered = [cat for cat in ground_truth[file_name] if cat.lower() != 'others']
                        pred_categories_filtered = [cat for cat in result['approaches'][approach_name]['categories'] if cat.lower() != 'others']
                        
                        all_gt.append(gt_categories_filtered)
                        all_pred.append(pred_categories_filtered)
            
            if all_gt and all_pred:
                metrics = self.metrics_calculator.calculate_batch_metrics(all_gt, all_pred)
                overall_metrics[approach_name] = metrics
        
        return overall_metrics
    
    def _save_results_excel(self, results: Dict, excel_path: Path):
        """Save results to Excel file with multiple sheets, formatted and sorted."""
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = []
            for result in results['results']:
                if result['status'] == 'completed':
                    row = {
                        'File Name': result['file_name'],
                        'Subject': result.get('subject', ''),
                        'Body': result.get('body', ''),
                        'Processing Time (s)': f"{result['processing_time']:.2f}",
                        'Ground Truth': ', '.join(result.get('ground_truth', [])),
                    }
                    
                    # Add approach results
                    for approach_name, approach_data in result['approaches'].items():
                        approach_title = approach_name.replace("_", " ").title()
                        
                        # Categories and confidence scores
                        categories_with_confidence = []
                        for cat_detail in approach_data.get('category_details', []):
                            cat = cat_detail.get('category', '')
                            conf = cat_detail.get('confidence', 0.0)
                            categories_with_confidence.append(f"{cat} ({conf:.3f})")
                        
                        row['Predicted Categories'] = ', '.join(approach_data['categories'])
                        row['Categories with Confidence'] = ', '.join(categories_with_confidence)
                        row['RAG Time (s)'] = f"{approach_data['processing_time']:.2f}"
                        
                        # Add metrics if available
                        if 'metrics' in approach_data:
                            metrics = approach_data['metrics']
                            row['Precision'] = round(metrics.get('precision', 0), 2) if metrics.get('precision') is not None else None
                            row['Recall'] = round(metrics.get('recall', 0), 2) if metrics.get('recall') is not None else None
                            row['F1'] = round(metrics.get('f1_score', 0), 2) if metrics.get('f1_score') is not None else None
                            row['Exact Match'] = round(metrics.get('exact_match', 0), 2) if metrics.get('exact_match') is not None else None
                            row['True Positives'] = int(metrics.get('tp', 0)) if metrics.get('tp') is not None else None
                            row['False Positives'] = int(metrics.get('fp', 0)) if metrics.get('fp') is not None else None
                            row['False Negatives'] = int(metrics.get('fn', 0)) if metrics.get('fn') is not None else None
                            row['False Negatives List'] = ', '.join(metrics.get('missed_categories', [])) if metrics.get('missed_categories') else ''
                            row['Jaccard Similarity'] = round(metrics.get('jaccard_similarity', 0), 2) if metrics.get('jaccard_similarity') is not None else None
                    
                    summary_data.append(row)
                else:
                    summary_data.append({
                        'File Name': result['file_name'],
                        'Status': result['status'],
                        'Error': result.get('error', '')
                    })
            
            summary_df = pd.DataFrame(summary_data)
            # Sort by File Name
            summary_df = summary_df.sort_values('File Name')
            summary_df.to_excel(writer, sheet_name='Results', index=False)
            
            # Format the Results sheet with auto-filter and left-alignment
            worksheet = writer.sheets['Results']
            # Enable auto-filter
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Apply left alignment to all cells, but disable text wrapping for subject/body columns
            from openpyxl.styles import Alignment
            left_align = Alignment(horizontal='left', vertical='top', wrap_text=False)
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = left_align
            
            # Detailed results sheet with one row per category prediction
            detailed_data = []
            for result in results['results']:
                if result['status'] == 'completed':
                    base_info = {
                        'File Name': result['file_name'],
                        'Subject': result.get('subject', ''),
                        'Body': result.get('body', '')[:500] + '...' if len(result.get('body', '')) > 500 else result.get('body', ''),  # Truncate long bodies
                        'Ground Truth': ', '.join(result.get('ground_truth', []))
                    }
                    
                    # Add one row per approach per category
                    for approach_name, approach_data in result['approaches'].items():
                        approach_title = "RAG"  # Simplified approach name
                        
                        if approach_data.get('category_details'):
                            for cat_detail in approach_data['category_details']:
                                detailed_row = base_info.copy()
                                
                                # Get evidence from RAG lookup as justification, truncate if too long
                                evidence = cat_detail.get('evidence', 'No supporting evidence found in RAG lookup')
                                # Truncate long evidence and show only the first sentence or first 150 characters
                                if evidence and len(evidence) > 150:
                                    sentences = evidence.split('. ')
                                    justification = sentences[0] + ('.' if len(sentences[0]) < len(evidence) else '') + '...'
                                else:
                                    justification = evidence
                                
                                # Get issue types that led to this category
                                issue_types = cat_detail.get('issue_types', [])
                                issue_types_str = ', '.join(issue_types) if issue_types else 'No issue types found'
                                
                                detailed_row.update({
                                    'Approach': approach_title,
                                    'Predicted Category': cat_detail.get('category', ''),
                                    'Confidence Score': cat_detail.get('confidence', 0.0),
                                    'Issue Types': issue_types_str,
                                    'Justification': justification,
                                    'Processing Time (s)': f"{approach_data['processing_time']:.2f}"
                                })
                                detailed_data.append(detailed_row)
                        else:
                            # No categories found
                            detailed_row = base_info.copy()
                            detailed_row.update({
                                'Approach': approach_title,
                                'Predicted Category': 'No categories found',
                                'Confidence Score': 0.0,
                                'Issue Types': 'N/A',
                                'Justification': 'No categories were predicted above the confidence threshold',
                                'Processing Time (s)': f"{approach_data['processing_time']:.2f}"
                            })
                            detailed_data.append(detailed_row)
            
            if detailed_data:
                detailed_df = pd.DataFrame(detailed_data)
                # Sort by File Name
                detailed_df = detailed_df.sort_values('File Name')
                detailed_df.to_excel(writer, sheet_name='Detailed Results', index=False)
                
                # Format the Detailed Results sheet with auto-filter and left-alignment
                detailed_worksheet = writer.sheets['Detailed Results']
                # Enable auto-filter
                detailed_worksheet.auto_filter.ref = detailed_worksheet.dimensions
                
                # Apply left alignment to all cells
                for row in detailed_worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = left_align
            
            # Overall metrics sheet if available
            if 'overall_metrics' in results:
                metrics_data = []
                for approach_name, metrics in results['overall_metrics'].items():
                    metrics_data.append({
                        'Approach': approach_name.replace('_', ' ').title(),
                        'Total Files': int(metrics.get('total_files', 0)) if metrics.get('total_files') is not None else None,
                        'Micro Precision': round(metrics.get('micro_precision', 0), 2) if metrics.get('micro_precision') is not None else None,
                        'Micro Recall': round(metrics.get('micro_recall', 0), 2) if metrics.get('micro_recall') is not None else None,
                        'Micro F1 Score': round(metrics.get('micro_f1', 0), 2) if metrics.get('micro_f1') is not None else None,
                        'Macro Precision': round(metrics.get('macro_precision', 0), 2) if metrics.get('macro_precision') is not None else None,
                        'Macro Recall': round(metrics.get('macro_recall', 0), 2) if metrics.get('macro_recall') is not None else None,
                        'Macro F1 Score': round(metrics.get('macro_f1', 0), 2) if metrics.get('macro_f1') is not None else None,
                        'Exact Match Accuracy': round(metrics.get('exact_match_accuracy', 0), 2) if metrics.get('exact_match_accuracy') is not None else None,
                        'Average Jaccard Similarity': round(metrics.get('average_jaccard_similarity', 0), 2) if metrics.get('average_jaccard_similarity') is not None else None,
                        'Perfect Predictions': int(metrics.get('perfect_predictions', 0)) if metrics.get('perfect_predictions') is not None else None,
                        'Total True Positives': int(metrics.get('total_tp', 0)) if metrics.get('total_tp') is not None else None,
                        'Total False Positives': int(metrics.get('total_fp', 0)) if metrics.get('total_fp') is not None else None,
                        'Total False Negatives': int(metrics.get('total_fn', 0)) if metrics.get('total_fn') is not None else None
                    })
                
                metrics_df = pd.DataFrame(metrics_data)
                metrics_df.to_excel(writer, sheet_name='Overall Metrics', index=False)
                
                # Format the Overall Metrics sheet with left-alignment
                metrics_worksheet = writer.sheets['Overall Metrics']
                for row in metrics_worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = left_align
            
            # Processing stats sheet
            stats_data = [{
                'Metric': k.replace('_', ' ').title(),
                'Value': str(v)
            } for k, v in results['processing_stats'].items()]
            
            stats_df = pd.DataFrame(stats_data)
            stats_df.to_excel(writer, sheet_name='Processing Stats', index=False)
            
            # Format the Processing Stats sheet with left-alignment
            stats_worksheet = writer.sheets['Processing Stats']
            for row in stats_worksheet.iter_rows():
                for cell in row:
                    cell.alignment = left_align


# Convenience functions for easy usage
def process_lot_pdfs(pdf_folder: str, 
                    ground_truth_file: str = None,
                    enable_llm: bool = False,
                    enable_metrics: bool = True,
                    output_folder: str = None) -> Dict:
    """
    Convenience function to process a lot of PDFs with simple parameters.
    
    Args:
        pdf_folder: Path to folder containing PDF files
        ground_truth_file: Optional path to ground truth Excel file
        enable_llm: Whether to enable Pure LLM approach (default: False)
        enable_metrics: Whether to enable metrics calculation (default: True)
        output_folder: Optional custom output folder
        
    Returns:
        Dictionary containing processing results and metrics
    """
    # Create temporary batch config
    temp_config = {
        'batch_processing': {
            'enabled': True,
            'approaches': {
                'hybrid_rag': {'enabled': True, 'priority': 1},
                'pure_llm': {'enabled': enable_llm, 'priority': 2}
            },
            'evaluation': {
                'enabled': enable_metrics,
                'auto_detect_ground_truth': ground_truth_file is None,
                'ground_truth_patterns': ["EDMS*.xlsx", "LOT-*.xlsx", "ground_truth*.xlsx", "*_labels.xlsx"]
            },
            'output': {
                'results_folder': output_folder or 'results',
                'save_format': 'xlsx'
            },
            'processing': {
                'max_pages_per_pdf': 2,
                'skip_on_error': True,
                'rate_limit_delay': 3
            }
        }
    }
    
    # Save temporary config
    temp_config_path = Path("temp_batch_config.yaml")
    with open(temp_config_path, 'w') as f:
        yaml.dump(temp_config, f)
    
    try:
        # Process with temporary config
        processor = BatchPDFProcessor(batch_config_path=str(temp_config_path))
        return processor.process_pdf_folder(pdf_folder, ground_truth_file, output_folder)
    finally:
        # Cleanup temporary config
        if temp_config_path.exists():
            temp_config_path.unlink()


if __name__ == "__main__":
    # Example usage
    import argparse
    
    parser = argparse.ArgumentParser(description='Batch PDF Processing')
    parser.add_argument('pdf_folder', help='Path to PDF folder')
    parser.add_argument('--ground-truth', help='Path to ground truth Excel file')
    parser.add_argument('--enable-llm', action='store_true', help='Enable Pure LLM approach')
    parser.add_argument('--disable-metrics', action='store_true', help='Disable metrics calculation')
    parser.add_argument('--output', help='Output folder path')
    
    args = parser.parse_args()
    
    # Process PDFs
    results = process_lot_pdfs(
        pdf_folder=args.pdf_folder,
        ground_truth_file=args.ground_truth,
        enable_llm=args.enable_llm,
        enable_metrics=not args.disable_metrics,
        output_folder=args.output
    )
    
    print(f"Processing completed! Results saved to: {results.get('output_path', 'results/')}")